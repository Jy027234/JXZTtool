from __future__ import annotations

import importlib.util
import os
from pathlib import Path
import unittest

from parsecore.bootstrap import build_runtime
from parsecore.models import BlockType, ParseRequest, SemanticRole
from tests.support import TemporaryWorkspace


EXCEL_CONFIG = """
[project]
name = "test-excel"
mode = "embedded-sdk"

[runtime]
execution_mode = "inline"
max_workers = 2
poll_interval_ms = 25

[storage]
database_url = "__DB_URL__"
object_store = "local://./var/uploads"

[index]
mode = "hybrid"

[translation]
enabled = true
strategy = "lazy"

[product]
adapter = "embedded"

[[parsers]]
name = "excel-native"
media_types = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
    "application/vnd.ms-excel",
]
extensions = [".xlsx", ".xlsm", ".xls"]
""".strip()


EXCEL_VISIBLE_ONLY_CONFIG = EXCEL_CONFIG + """

[parsers.options]
include_hidden_sheets = false
max_rows_per_sheet = 2
max_cols_per_sheet = 2
""".rstrip()


EXCEL_METADATA_LIMIT_CONFIG = EXCEL_CONFIG + """

[parsers.options]
max_metadata_cells = 4
""".rstrip()


EXCEL_SAMPLE_DIR = Path(os.environ.get("PARSECORE_EXCEL_SAMPLE_DIR", "D:/app/uploads"))


@unittest.skipUnless(importlib.util.find_spec("openpyxl"), "openpyxl not installed")
class ExcelParserTests(unittest.TestCase):
    def _create_workbook(self, workspace: TemporaryWorkspace, name: str):
        from openpyxl import Workbook

        assert workspace.root is not None
        target = workspace.root / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BOM"
        sheet.append(["Part", "Qty", "Cost"])
        sheet.append(["A-100", 2, 3.5])
        sheet.append(["Total", "", "=B2*C2"])
        hidden = workbook.create_sheet("HiddenCalc")
        hidden.sheet_state = "hidden"
        hidden.append(["Secret", "Value"])
        hidden.append(["Internal", 42])
        workbook.save(target)
        return target

    def _create_titled_workbook(self, workspace: TemporaryWorkspace, name: str):
        from openpyxl import Workbook

        assert workspace.root is not None
        target = workspace.root / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Plan"
        sheet.append(["Maintenance Plan", ""])
        sheet.merge_cells("A1:B1")
        sheet.append(["Task", "Owner"])
        sheet.append(["Inspect", "QA"])
        workbook.save(target)
        return target

    def _create_large_metadata_workbook(self, workspace: TemporaryWorkspace, name: str):
        from openpyxl import Workbook

        assert workspace.root is not None
        target = workspace.root / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Large"
        sheet.append(["A", "B", "C"])
        sheet.append(["1", "2", "3"])
        sheet.append(["4", "5", "6"])
        workbook.save(target)
        return target

    def _create_multi_table_workbook(self, workspace: TemporaryWorkspace, name: str):
        from openpyxl import Workbook

        assert workspace.root is not None
        target = workspace.root / name
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Mixed"
        sheet.append(["Part", "Qty"])
        sheet.append(["A-100", 2])
        sheet.append([])
        sheet.append(["Checklist", "Status"])
        sheet.append(["Torque", "Done"])
        workbook.save(target)
        return target

    def test_excel_parser_emits_sheet_table_blocks(self) -> None:
        with TemporaryWorkspace(EXCEL_CONFIG) as workspace:
            workbook_path = self._create_workbook(workspace, "bom.xlsx")
            runtime = build_runtime(workspace.config_path)
            outcome = runtime.submit(
                ParseRequest(
                    doc_id="doc-excel",
                    file_path=str(workbook_path),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            )

        self.assertEqual(outcome.blocks[0].type, BlockType.TITLE)
        self.assertEqual(outcome.blocks[0].metadata["parser"], "excel-native")
        table_blocks = [block for block in outcome.blocks if block.type == BlockType.TABLE]
        self.assertEqual(len(table_blocks), 2)
        bom = table_blocks[0]
        self.assertIn("| Part | Qty | Cost |", bom.content)
        self.assertIn("| A-100 | 2 | 3.5 |", bom.content)
        self.assertEqual(bom.metadata["semantic_role"], SemanticRole.TABLE.value)
        self.assertEqual(bom.metadata["table_type"], "spreadsheet")
        self.assertEqual(bom.metadata["sheet_name"], "BOM")
        self.assertEqual(bom.metadata["cell_range"], "A1:C3")
        self.assertEqual(bom.metadata["sheet_table_index"], 1)
        self.assertEqual(bom.metadata["sheet_table_count"], 1)
        self.assertEqual(bom.metadata["row_range"], "1:3")
        self.assertEqual(bom.metadata["column_range"], "A:C")
        self.assertEqual(bom.metadata["rows"], 3)
        self.assertEqual(bom.metadata["cols"], 3)
        self.assertTrue(bom.metadata["has_formula"])
        self.assertEqual(bom.metadata["formula_count"], 1)
        self.assertFalse(bom.metadata["hidden_sheet"])
        hidden = table_blocks[1]
        self.assertEqual(hidden.metadata["sheet_name"], "HiddenCalc")
        self.assertTrue(hidden.metadata["hidden_sheet"])

    def test_excel_parser_prefers_xlsx_suffix_over_generic_excel_mime(self) -> None:
        with TemporaryWorkspace(EXCEL_CONFIG) as workspace:
            workbook_path = self._create_workbook(workspace, "generic-mime.xlsx")
            runtime = build_runtime(workspace.config_path)
            outcome = runtime.submit(
                ParseRequest(
                    doc_id="doc-excel-generic-mime",
                    file_path=str(workbook_path),
                    media_type="application/vnd.ms-excel",
                )
            )

        table_blocks = [block for block in outcome.blocks if block.type == BlockType.TABLE]
        self.assertEqual(len(table_blocks), 2)
        self.assertEqual(table_blocks[0].metadata["sheet_name"], "BOM")
        self.assertIn("| Part | Qty | Cost |", table_blocks[0].content)

    def test_excel_parser_splits_tables_separated_by_blank_rows(self) -> None:
        with TemporaryWorkspace(EXCEL_CONFIG) as workspace:
            workbook_path = self._create_multi_table_workbook(workspace, "mixed.xlsx")
            runtime = build_runtime(workspace.config_path)
            outcome = runtime.submit(
                ParseRequest(
                    doc_id="doc-excel-mixed",
                    file_path=str(workbook_path),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            )

        table_blocks = [block for block in outcome.blocks if block.type == BlockType.TABLE]
        self.assertEqual(len(table_blocks), 2)
        first, second = table_blocks
        self.assertEqual(first.metadata["sheet_name"], "Mixed")
        self.assertEqual(first.metadata["sheet_table_index"], 1)
        self.assertEqual(first.metadata["sheet_table_count"], 2)
        self.assertEqual(first.metadata["cell_range"], "A1:B2")
        self.assertIn("| Part | Qty |", first.content)
        self.assertEqual(second.metadata["sheet_table_index"], 2)
        self.assertEqual(second.metadata["sheet_table_count"], 2)
        self.assertEqual(second.metadata["cell_range"], "A4:B5")
        self.assertIn("| Checklist | Status |", second.content)

    def test_excel_parser_honors_visibility_and_size_options(self) -> None:
        with TemporaryWorkspace(EXCEL_VISIBLE_ONLY_CONFIG) as workspace:
            workbook_path = self._create_workbook(workspace, "limited.xlsx")
            runtime = build_runtime(workspace.config_path)
            outcome = runtime.submit(
                ParseRequest(
                    doc_id="doc-excel-limited",
                    file_path=str(workbook_path),
                    media_type=None,
                )
            )

        table_blocks = [block for block in outcome.blocks if block.type == BlockType.TABLE]
        self.assertEqual(len(table_blocks), 1)
        table = table_blocks[0]
        self.assertEqual(table.metadata["sheet_name"], "BOM")
        self.assertEqual(table.metadata["cell_range"], "A1:B2")
        self.assertEqual(table.metadata["rows"], 2)
        self.assertEqual(table.metadata["cols"], 2)
        self.assertTrue(table.metadata["truncated"])
        self.assertNotIn("HiddenCalc", {block.metadata.get("sheet_name") for block in table_blocks})

    def test_excel_parser_detects_title_header_and_merged_cells(self) -> None:
        with TemporaryWorkspace(EXCEL_CONFIG) as workspace:
            workbook_path = self._create_titled_workbook(workspace, "titled.xlsx")
            runtime = build_runtime(workspace.config_path)
            outcome = runtime.submit(
                ParseRequest(
                    doc_id="doc-excel-titled",
                    file_path=str(workbook_path),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            )

        table_blocks = [block for block in outcome.blocks if block.type == BlockType.TABLE]
        self.assertEqual(len(table_blocks), 1)
        table = table_blocks[0]
        self.assertEqual(table.metadata["table_title"], "Maintenance Plan")
        self.assertEqual(table.metadata["title_row"], 1)
        self.assertEqual(table.metadata["header_row"], 2)
        self.assertEqual(table.metadata["header_values"], ["Task", "Owner"])
        self.assertEqual(table.metadata["cell_range"], "A2:B3")
        self.assertEqual(table.metadata["source_cell_range"], "A1:B3")
        self.assertEqual(table.metadata["merged_cells"], ["A1:B1"])
        self.assertIn("| Task | Owner |", table.content)

    def test_excel_parser_limits_large_cell_metadata(self) -> None:
        with TemporaryWorkspace(EXCEL_METADATA_LIMIT_CONFIG) as workspace:
            workbook_path = self._create_large_metadata_workbook(workspace, "large.xlsx")
            runtime = build_runtime(workspace.config_path)
            outcome = runtime.submit(
                ParseRequest(
                    doc_id="doc-excel-large",
                    file_path=str(workbook_path),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            )

        table_blocks = [block for block in outcome.blocks if block.type == BlockType.TABLE]
        self.assertEqual(len(table_blocks), 1)
        table = table_blocks[0]
        self.assertNotIn("cells", table.metadata)
        self.assertEqual(table.metadata["cells_preview"], [["A", "B", "C"]])
        self.assertTrue(table.metadata["cells_truncated"])
        self.assertEqual(table.metadata["cells_total"], 9)
        self.assertEqual(table.metadata["cells_preview_rows"], 1)

    def test_excel_parser_handles_real_upload_samples(self) -> None:
        if not EXCEL_SAMPLE_DIR.exists():
            self.skipTest(f"{EXCEL_SAMPLE_DIR} not available")
        samples = sorted(
            path
            for path in EXCEL_SAMPLE_DIR.iterdir()
            if path.suffix.lower() in {".xls", ".xlsx", ".xlsm"}
        )
        if not samples:
            self.skipTest(f"no spreadsheet samples found in {EXCEL_SAMPLE_DIR}")
        if any(path.suffix.lower() == ".xls" for path in samples) and not importlib.util.find_spec("xlrd"):
            self.skipTest("xlrd not installed")

        media_types = {
            ".xls": "application/vnd.ms-excel",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
        }
        with TemporaryWorkspace(EXCEL_CONFIG) as workspace:
            runtime = build_runtime(workspace.config_path)
            results = {
                sample.name: runtime.submit(
                    ParseRequest(
                        doc_id=f"sample-{index}",
                        file_path=str(sample),
                        media_type=media_types[sample.suffix.lower()],
                    )
                )
                for index, sample in enumerate(samples, start=1)
            }

        self.assertEqual(len(results), len(samples))
        xls_seen = False
        xlsx_metadata_seen = False
        for sample in samples:
            outcome = results[sample.name]
            table_blocks = [block for block in outcome.blocks if block.type == BlockType.TABLE]
            self.assertGreaterEqual(len(table_blocks), 1, sample.name)
            self.assertTrue(any(block.content.strip() for block in table_blocks), sample.name)
            for block in table_blocks:
                self.assertEqual(block.metadata["parser"], "excel-native", sample.name)
                self.assertIn("sheet_name", block.metadata, sample.name)
                self.assertIn("cell_range", block.metadata, sample.name)
                self.assertIn("source_cell_range", block.metadata, sample.name)
                self.assertIn("header_row", block.metadata, sample.name)
                self.assertIn("header_values", block.metadata, sample.name)
            if sample.suffix.lower() == ".xls":
                xls_seen = True
            else:
                xlsx_metadata_seen = xlsx_metadata_seen or any(
                    block.metadata.get("merged_cells") or block.metadata.get("table_title")
                    for block in table_blocks
                )

        self.assertTrue(xls_seen)
        self.assertTrue(xlsx_metadata_seen)


if __name__ == "__main__":
    unittest.main()
