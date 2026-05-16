from __future__ import annotations

from dataclasses import dataclass
import re
import time
from pathlib import Path
from typing import Any, Callable, Mapping, Sequence
import xml.etree.ElementTree as ET
import zipfile

from .config import OcrProviderSettings
from .contracts import ParserAdapter
from .garble import detect_page_garble_reason
from .models import Block, BlockType, ParseRequest, SemanticRole
from .ocr import OcrConfigurationError, OcrRequestError, build_ocr_engine
from .ocr_cache import PageOcrCache, get_default_cache
from .stubs import StubParser


_DEFAULT_OCR_PROVIDER_SETTINGS = OcrProviderSettings(
    enabled=True,
    provider="rapidocr",
)

_DOCX_NOISE_ROLES = {
    SemanticRole.HEADER_FOOTER.value,
    SemanticRole.PARSE_ARTIFACT.value,
    SemanticRole.VERSION_CELL.value,
    SemanticRole.PAGE_REF_CELL.value,
}

_DOCX_ROMAN_PAGE_PATTERN = r"[IVXLCDMⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]{1,10}(?:\s*至\s*[IVXLCDMⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]{1,10})?"
_DOCX_PAGE_REF_CORE_PATTERN = (
    r"(?:[A-Z]?\d+(?:[-./]\d+)+|"
    + _DOCX_ROMAN_PAGE_PATTERN
    + r"|[一二三四五六七八九十百千]+/\d+)"
)
_DOCX_PAGE_REF_PATTERN = re.compile(
    rf"^\s*(?:页码\s*[:：]\s*)?(?P<label>{_DOCX_PAGE_REF_CORE_PATTERN})\s*$",
    re.IGNORECASE,
)
_DOCX_TRAILING_PAGE_REF_PATTERN = re.compile(
    rf"(?P<label>{_DOCX_PAGE_REF_CORE_PATTERN})\s*$",
    re.IGNORECASE,
)
_DOCX_VERSION_PATTERN = re.compile(
    r"^\s*(?:R\d+(?:TR?\d+)?|Rev(?:ision)?\s*[A-Z0-9.-]+|版次(?:/修订)?|修订(?:记录)?|版本(?:号)?|版号)\s*$",
    re.IGNORECASE,
)
_DOCX_DATE_PATTERN = re.compile(r"^\s*\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\s*$")
_DOCX_TMP_ARTIFACT_PATTERN = re.compile(r"^\s*tmp[\\/_ .-].*", re.IGNORECASE)
_DOCX_NUMBERED_HEADING_PATTERN = re.compile(
    r"^\s*(?P<number>(?:\d+(?:\.\d+)+|\d+))(?:[.、)])?\s*(?P<title>.+?)\s*$"
)
_DOCX_CHAPTER_HEADING_PATTERN = re.compile(
    r"^\s*(?P<number>第[\d一二三四五六七八九十百千]+[章节条款])\s*(?P<title>.+?)\s*$"
)
_DOCX_APPENDIX_HEADING_PATTERN = re.compile(
    r"^\s*(?P<number>(?:附录|附件|Appendix|Annex)\s*[A-Z0-9一二三四五六七八九十]*)\s*(?P<title>.*)$",
    re.IGNORECASE,
)
_DOCX_HEADING_STYLE_PATTERN = re.compile(r"(?:heading|标题)", re.IGNORECASE)
_DOCX_TOC_HEADING_PATTERN = re.compile(r"^\s*(?:目录|目次|CONTENTS?|TABLE OF CONTENTS)\s*$", re.IGNORECASE)
_DOCX_LEP_HEADING_PATTERN = re.compile(
    r"(?:有效页清单|LIST OF EFFECTIVE PAGES)",
    re.IGNORECASE,
)
_DOCX_REVISION_HEADING_PATTERN = re.compile(
    r"(?:修订记录|修订页|版次表|版次/修订|RECORD OF REVISIONS|REVISION(?:\s+RECORD|\s+HISTORY)?)",
    re.IGNORECASE,
)
_DOCX_DISTRIBUTION_HEADING_PATTERN = re.compile(
    r"(?:分发清单|放行人员清单|外委单位清单|DISTRIBUTION LIST|AUTHORIZED RELEASE|VENDOR LIST)",
    re.IGNORECASE,
)
_DOCX_SIGNATURE_HEADING_PATTERN = re.compile(
    r"(?:签字|签署|批准|审批|SIGNATURE|APPROVAL)",
    re.IGNORECASE,
)
_DOCX_HEADER_FOOTER_PATTERN = re.compile(
    r"^\s*(?:页码\s*[:：]\s*)?(?:第\s*\d+\s*页|Page\s+\d+|[A-Z]{2,}\s+\d+(?:\.\d+)*)\s*$",
    re.IGNORECASE,
)
_OCR_TOKEN_RE = re.compile(r"\w+", re.UNICODE)


@dataclass(slots=True)
class _DocxParseState:
    body_started: bool = False
    toc_active: bool = False
    lep_active: bool = False
    last_heading_text: str | None = None
    last_heading_role: str | None = None
    last_heading_page_type: str = "body"
    heading_context_open: bool = False


@dataclass(slots=True, frozen=True)
class _DocxParagraphInfo:
    text: str
    style: str | None = None
    outline_level: int | None = None
    has_page_break: bool = False


@dataclass(slots=True, frozen=True)
class _DocxClassification:
    block_type: BlockType
    semantic_role: str
    content: str
    page_type: str
    logical_page_label: str | None = None
    heading_level: int | None = None
    normalized_title: str | None = None
    kind: str | None = None


def _xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _normalize_inline_whitespace(text: str) -> str:
    return re.sub(r"[ \t\u3000]+", " ", str(text or "")).strip()


def _docx_paragraph_info(
    paragraph: ET.Element,
    *,
    namespaces: Mapping[str, str],
) -> _DocxParagraphInfo:
    parts: list[str] = []
    has_page_break = False
    for node in paragraph.iter():
        local_name = _xml_local_name(node.tag)
        if local_name == "t" and node.text:
            parts.append(node.text)
        elif local_name == "tab":
            parts.append(" ")
        elif local_name in {"br", "cr"}:
            parts.append("\n")
            if local_name == "br":
                br_type = node.get(f"{{{namespaces['w']}}}type") or node.get("type", "")
                if br_type in {"page", "column"}:
                    has_page_break = True
        elif local_name == "lastRenderedPageBreak":
            has_page_break = True
    lines = [_normalize_inline_whitespace(item) for item in "".join(parts).splitlines()]
    text = "\n".join(item for item in lines if item).strip()

    style = None
    style_node = paragraph.find("./w:pPr/w:pStyle", namespaces)
    if style_node is not None:
        style = style_node.get(f"{{{namespaces['w']}}}val") or style_node.get("val")
    outline_level = None
    outline_node = paragraph.find("./w:pPr/w:outlineLvl", namespaces)
    if outline_node is not None:
        raw_level = outline_node.get(f"{{{namespaces['w']}}}val") or outline_node.get("val")
        try:
            outline_level = int(raw_level) + 1 if raw_level is not None else None
        except (TypeError, ValueError):
            outline_level = None
    return _DocxParagraphInfo(text=text, style=style, outline_level=outline_level, has_page_break=has_page_break)


def _docx_extract_page_label(text: str) -> str | None:
    stripped = _normalize_inline_whitespace(text)
    if not stripped:
        return None
    match = _DOCX_PAGE_REF_PATTERN.match(stripped)
    if match:
        return str(match.group("label") or "").strip() or None
    trailing = _DOCX_TRAILING_PAGE_REF_PATTERN.search(stripped)
    if trailing:
        label = str(trailing.group("label") or "").strip()
        if label and label != stripped:
            return label
    return None


def _docx_extract_numbered_heading(text: str) -> tuple[str, str, str | None, int | None] | None:
    stripped = _normalize_inline_whitespace(text)
    if not stripped:
        return None
    if _DOCX_PAGE_REF_PATTERN.match(stripped):
        return None
    logical_page = _docx_extract_page_label(stripped)
    if logical_page:
        trailing = _DOCX_TRAILING_PAGE_REF_PATTERN.search(stripped)
        if trailing is not None and trailing.start() > 0:
            stripped = stripped[: trailing.start()].rstrip()

    appendix_match = _DOCX_APPENDIX_HEADING_PATTERN.match(stripped)
    if appendix_match:
        number = _normalize_inline_whitespace(appendix_match.group("number") or "")
        title = _normalize_inline_whitespace(appendix_match.group("title") or "")
        normalized = number if not title else f"{number} {title}".strip()
        return normalized, SemanticRole.APPENDIX.value, logical_page, 1

    chapter_match = _DOCX_CHAPTER_HEADING_PATTERN.match(stripped)
    if chapter_match:
        number = _normalize_inline_whitespace(chapter_match.group("number") or "")
        title = _normalize_inline_whitespace(chapter_match.group("title") or "")
        normalized = number if not title else f"{number} {title}".strip()
        return normalized, SemanticRole.BODY_SECTION.value, logical_page, 1

    numbered_match = _DOCX_NUMBERED_HEADING_PATTERN.match(stripped)
    if numbered_match:
        number = _normalize_inline_whitespace(numbered_match.group("number") or "")
        title = _normalize_inline_whitespace(numbered_match.group("title") or "")
        if title:
            normalized = f"{number} {title}".strip()
            heading_level = number.count(".") + 1
            return normalized, SemanticRole.BODY_SECTION.value, logical_page, heading_level
    return None


def _is_docx_heading_style(style: str | None) -> bool:
    if not style:
        return False
    return _DOCX_HEADING_STYLE_PATTERN.search(style) is not None


def _classify_docx_noise(text: str) -> str | None:
    stripped = _normalize_inline_whitespace(text)
    if not stripped:
        return None
    if _DOCX_TMP_ARTIFACT_PATTERN.match(stripped):
        return SemanticRole.PARSE_ARTIFACT.value
    if _DOCX_VERSION_PATTERN.match(stripped):
        return SemanticRole.VERSION_CELL.value
    if _DOCX_PAGE_REF_PATTERN.match(stripped):
        return SemanticRole.PAGE_REF_CELL.value
    if _DOCX_HEADER_FOOTER_PATTERN.match(stripped) or _DOCX_DATE_PATTERN.match(stripped):
        return SemanticRole.HEADER_FOOTER.value
    return None


def _classify_docx_paragraph(
    info: _DocxParagraphInfo,
    *,
    state: _DocxParseState,
) -> _DocxClassification:
    stripped = info.text.strip()
    if not stripped:
        return _DocxClassification(
            block_type=BlockType.PARAGRAPH,
            semantic_role=SemanticRole.PARAGRAPH.value,
            content="",
            page_type="body",
            kind=BlockType.PARAGRAPH.value,
        )

    if _DOCX_TOC_HEADING_PATTERN.match(stripped):
        state.toc_active = True
        state.lep_active = False
        state.heading_context_open = True
        return _DocxClassification(
            block_type=BlockType.TITLE,
            semantic_role=SemanticRole.FRONT_MATTER.value,
            content=stripped,
            page_type="front_matter",
            normalized_title=stripped,
            kind=BlockType.TITLE.value,
        )

    if _DOCX_LEP_HEADING_PATTERN.search(stripped):
        state.toc_active = False
        state.lep_active = True
        state.heading_context_open = True
        return _DocxClassification(
            block_type=BlockType.TITLE,
            semantic_role=SemanticRole.FRONT_MATTER.value,
            content=_normalize_inline_whitespace(stripped),
            page_type="front_matter",
            normalized_title=_normalize_inline_whitespace(stripped),
            kind=BlockType.TITLE.value,
        )

    if _DOCX_REVISION_HEADING_PATTERN.search(stripped):
        state.toc_active = False
        state.lep_active = False
        state.heading_context_open = True
        return _DocxClassification(
            block_type=BlockType.TITLE,
            semantic_role=SemanticRole.FRONT_MATTER.value,
            content=_normalize_inline_whitespace(stripped),
            page_type="front_matter",
            normalized_title=_normalize_inline_whitespace(stripped),
            kind=BlockType.TITLE.value,
        )

    if _DOCX_DISTRIBUTION_HEADING_PATTERN.search(stripped):
        state.toc_active = False
        state.lep_active = False
        state.heading_context_open = True
        return _DocxClassification(
            block_type=BlockType.TITLE,
            semantic_role=SemanticRole.FRONT_MATTER.value,
            content=_normalize_inline_whitespace(stripped),
            page_type="front_matter",
            normalized_title=_normalize_inline_whitespace(stripped),
            kind=BlockType.TITLE.value,
        )

    if _DOCX_SIGNATURE_HEADING_PATTERN.search(stripped) and not state.body_started:
        state.toc_active = False
        state.lep_active = False
        state.heading_context_open = True
        return _DocxClassification(
            block_type=BlockType.TITLE,
            semantic_role=SemanticRole.FRONT_MATTER.value,
            content=_normalize_inline_whitespace(stripped),
            page_type="signature",
            normalized_title=_normalize_inline_whitespace(stripped),
            kind=BlockType.TITLE.value,
        )

    heading = _docx_extract_numbered_heading(stripped)
    if heading is not None:
        normalized_title, inferred_role, logical_page_label, heading_level = heading
        if state.body_started or _is_docx_heading_style(info.style) or inferred_role == SemanticRole.APPENDIX.value or logical_page_label is None:
            page_type = "appendix" if inferred_role == SemanticRole.APPENDIX.value else "body"
            state.body_started = inferred_role != SemanticRole.FRONT_MATTER.value
            state.toc_active = False
            state.lep_active = False
            state.heading_context_open = True
            return _DocxClassification(
                block_type=BlockType.TITLE,
                semantic_role=inferred_role,
                content=normalized_title,
                page_type=page_type,
                logical_page_label=logical_page_label,
                heading_level=info.outline_level or heading_level,
                normalized_title=normalized_title,
                kind=BlockType.TITLE.value,
            )
        state.heading_context_open = False
        return _DocxClassification(
            block_type=BlockType.PARAGRAPH,
            semantic_role=SemanticRole.TOC_ENTRY.value,
            content=normalized_title,
            page_type="toc",
            logical_page_label=logical_page_label,
            normalized_title=normalized_title,
            kind=BlockType.PARAGRAPH.value,
        )

    noise_role = _classify_docx_noise(stripped)
    if noise_role is not None:
        state.heading_context_open = False
        return _DocxClassification(
            block_type=BlockType.PARAGRAPH,
            semantic_role=noise_role,
            content=_normalize_inline_whitespace(stripped),
            page_type="body" if state.body_started else "front_matter",
            logical_page_label=_docx_extract_page_label(stripped),
            kind=BlockType.PARAGRAPH.value,
        )

    role = _infer_semantic_role(stripped)
    if role == SemanticRole.PARAGRAPH.value and state.toc_active and _docx_extract_page_label(stripped):
        role = SemanticRole.TOC_ENTRY.value
    elif role == SemanticRole.PARAGRAPH.value and state.lep_active:
        role = SemanticRole.LEP_ENTRY.value

    page_type = "body"
    if role == SemanticRole.TOC_ENTRY.value:
        page_type = "toc"
    elif role == SemanticRole.LEP_ENTRY.value or not state.body_started:
        page_type = "front_matter"
    elif state.body_started and _DOCX_APPENDIX_HEADING_PATTERN.match(stripped):
        page_type = "appendix"

    if _is_docx_heading_style(info.style) and role == SemanticRole.PARAGRAPH.value:
        state.body_started = True
        state.heading_context_open = True
        normalized_title = _normalize_inline_whitespace(stripped)
        return _DocxClassification(
            block_type=BlockType.TITLE,
            semantic_role=SemanticRole.BODY_SECTION.value,
            content=normalized_title,
            page_type="body",
            logical_page_label=_docx_extract_page_label(stripped),
            heading_level=info.outline_level,
            normalized_title=normalized_title,
            kind=BlockType.TITLE.value,
        )

    state.heading_context_open = False
    return _DocxClassification(
        block_type=BlockType.PARAGRAPH,
        semantic_role=role if state.body_started or role in {SemanticRole.TOC_ENTRY.value, SemanticRole.LEP_ENTRY.value, SemanticRole.NOTE.value, SemanticRole.WARNING.value, SemanticRole.CAUTION.value} else SemanticRole.PARAGRAPH.value,
        content=stripped,
        page_type=page_type,
        logical_page_label=_docx_extract_page_label(stripped),
        kind=BlockType.PARAGRAPH.value,
    )


def _docx_table_cells(
    table: ET.Element,
    *,
    namespaces: Mapping[str, str],
) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table.findall("./w:tr", namespaces):
        values: list[str] = []
        for cell in row.findall("./w:tc", namespaces):
            parts: list[str] = []
            for paragraph in cell.findall("./w:p", namespaces):
                info = _docx_paragraph_info(paragraph, namespaces=namespaces)
                if info.text:
                    parts.append(info.text)
            values.append("\n".join(item for item in parts if item).strip())
        if any(value for value in values):
            rows.append(values)
    return rows


def _escape_markdown_table_cell(text: str) -> str:
    return str(text or "").replace("|", "\\|").replace("\n", "<br>").strip()


def _render_docx_table_markdown(rows: Sequence[Sequence[str]], *, header_rows: int = 1) -> str:
    normalized_rows = [list(row) for row in rows if any(str(cell or "").strip() for cell in row)]
    if not normalized_rows:
        return ""
    width = max((len(row) for row in normalized_rows), default=0)
    padded = [row + [""] * (width - len(row)) for row in normalized_rows]
    if width == 0:
        return ""
    rendered: list[str] = []
    header = padded[0]
    rendered.append("| " + " | ".join(_escape_markdown_table_cell(cell) for cell in header) + " |")
    rendered.append("| " + " | ".join("---" for _ in header) + " |")
    for row in padded[header_rows:]:
        rendered.append("| " + " | ".join(_escape_markdown_table_cell(cell) for cell in row) + " |")
    return "\n".join(rendered)


def _normalize_excel_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    isoformat = getattr(value, "isoformat", None)
    if callable(isoformat):
        return str(isoformat())
    return str(value).strip()


def _excel_column_letter(index: int) -> str:
    value = max(1, int(index))
    letters: list[str] = []
    while value:
        value, remainder = divmod(value - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def _excel_cell_ref(row: int, col: int) -> str:
    return f"{_excel_column_letter(col)}{row}"


def _trim_excel_rows(rows: Sequence[Sequence[str]]) -> tuple[list[list[str]], int, int]:
    non_empty_row_indexes = [
        index for index, row in enumerate(rows) if any(str(cell or "").strip() for cell in row)
    ]
    if not non_empty_row_indexes:
        return [], 0, 0
    first_row = non_empty_row_indexes[0]
    last_row = non_empty_row_indexes[-1]
    width = max((len(row) for row in rows), default=0)
    non_empty_col_indexes = [
        index
        for index in range(width)
        if any(
            index < len(row) and str(row[index] or "").strip()
            for row in rows[first_row : last_row + 1]
        )
    ]
    if not non_empty_col_indexes:
        return [], 0, 0
    first_col = non_empty_col_indexes[0]
    last_col = non_empty_col_indexes[-1]
    trimmed = [
        list(row[first_col : last_col + 1])
        for row in rows[first_row : last_row + 1]
    ]
    return trimmed, first_row + 1, first_col + 1


def _split_excel_table_regions(rows: Sequence[Sequence[str]]) -> list[tuple[list[list[str]], int]]:
    regions: list[tuple[list[list[str]], int]] = []
    current: list[list[str]] = []
    start_row = 0
    normalized_rows = [[str(cell or "").strip() for cell in row] for row in rows]
    for offset, row in enumerate(normalized_rows):
        index = offset + 1
        normalized = [str(cell or "").strip() for cell in row]
        next_row = normalized_rows[offset + 1] if offset + 1 < len(normalized_rows) else []
        starts_new_titled_region = (
            bool(current)
            and _excel_non_empty_count(normalized) == 1
            and _excel_non_empty_count(next_row) >= 2
        )
        if any(normalized) and starts_new_titled_region:
            regions.append((current, start_row))
            current = [normalized]
            start_row = index
            continue
        if any(normalized):
            if not current:
                start_row = index
            current.append(normalized)
            continue
        if current:
            regions.append((current, start_row))
            current = []
            start_row = 0
    if current:
        regions.append((current, start_row))
    return regions


def _excel_non_empty_count(row: Sequence[str]) -> int:
    return sum(1 for cell in row if str(cell or "").strip())


def _detect_excel_table_layout(
    rows: Sequence[Sequence[str]],
    *,
    start_row: int,
) -> tuple[list[list[str]], int, str | None, int | None, int | None]:
    normalized_rows = [[str(cell or "").strip() for cell in row] for row in rows]
    if (
        len(normalized_rows) >= 2
        and _excel_non_empty_count(normalized_rows[0]) == 1
        and _excel_non_empty_count(normalized_rows[1]) >= 2
    ):
        title = next(cell for cell in normalized_rows[0] if cell)
        header_row = start_row + 1
        return normalized_rows[1:], header_row, title, start_row, header_row
    header_row = start_row if normalized_rows else None
    return normalized_rows, start_row, None, None, header_row


def _excel_range_ref(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    return f"{_excel_cell_ref(start_row, start_col)}:{_excel_cell_ref(end_row, end_col)}"


def _excel_range_intersects(
    left: tuple[int, int, int, int],
    right: tuple[int, int, int, int],
) -> bool:
    left_min_row, left_min_col, left_max_row, left_max_col = left
    right_min_row, right_min_col, right_max_row, right_max_col = right
    return not (
        left_max_row < right_min_row
        or right_max_row < left_min_row
        or left_max_col < right_min_col
        or right_max_col < left_min_col
    )


def _excel_cells_metadata(rows: Sequence[Sequence[str]], *, max_cells: int) -> dict[str, Any]:
    row_count = len(rows)
    col_count = max((len(row) for row in rows), default=0)
    total_cells = row_count * col_count
    if total_cells <= max(0, int(max_cells)):
        return {
            "cells": [list(row) for row in rows],
            "cells_truncated": False,
            "cells_total": total_cells,
        }
    preview_rows = max(1, min(row_count, max(1, int(max_cells)) // max(col_count, 1)))
    return {
        "cells_preview": [list(row) for row in rows[:preview_rows]],
        "cells_truncated": True,
        "cells_total": total_cells,
        "cells_preview_rows": preview_rows,
    }


def _render_excel_table_markdown(rows: Sequence[Sequence[str]]) -> str:
    return _render_docx_table_markdown(rows)


def _classify_docx_table(
    cells: Sequence[Sequence[str]],
    *,
    state: _DocxParseState,
) -> tuple[str, str, str, str | None, list[list[str]]]:
    flattened = "\n".join(
        _normalize_inline_whitespace(cell)
        for row in cells
        for cell in row
        if _normalize_inline_whitespace(cell)
    )
    cell_roles: list[list[str]] = []
    non_empty_roles: set[str] = set()
    logical_page_labels: list[str] = []
    for row in cells:
        row_roles: list[str] = []
        for cell in row:
            normalized = _normalize_inline_whitespace(cell)
            role = _classify_docx_noise(normalized) or SemanticRole.PARAGRAPH.value
            row_roles.append(role)
            if normalized:
                non_empty_roles.add(role)
            label = _docx_extract_page_label(normalized)
            if label:
                logical_page_labels.append(label)
        cell_roles.append(row_roles)

    table_type = "general_table"
    semantic_role = SemanticRole.TABLE.value
    page_type = "body" if state.body_started else "front_matter"
    if _DOCX_LEP_HEADING_PATTERN.search(flattened) or state.last_heading_role == SemanticRole.FRONT_MATTER.value and state.last_heading_text and _DOCX_LEP_HEADING_PATTERN.search(state.last_heading_text):
        semantic_role = SemanticRole.LEP_ENTRY.value
        table_type = "effective_page_list"
        page_type = "front_matter"
    elif _DOCX_REVISION_HEADING_PATTERN.search(flattened) or state.last_heading_text and _DOCX_REVISION_HEADING_PATTERN.search(state.last_heading_text):
        semantic_role = SemanticRole.REVISION_RECORD.value
        table_type = "revision_record"
        page_type = "front_matter"
    elif _DOCX_DISTRIBUTION_HEADING_PATTERN.search(flattened) or state.last_heading_text and _DOCX_DISTRIBUTION_HEADING_PATTERN.search(state.last_heading_text):
        semantic_role = SemanticRole.DISTRIBUTION_LIST.value
        page_type = "front_matter"
        if re.search(r"放行人员", flattened):
            table_type = "release_personnel_list"
        elif re.search(r"外委单位", flattened):
            table_type = "outsourced_vendor_list"
        else:
            table_type = "distribution_list"
    elif state.last_heading_role == SemanticRole.APPENDIX.value:
        semantic_role = SemanticRole.APPENDIX.value
        page_type = "appendix"
        table_type = "appendix_table"
    elif non_empty_roles and non_empty_roles.issubset(_DOCX_NOISE_ROLES):
        semantic_role = SemanticRole.PARSE_ARTIFACT.value
        page_type = "front_matter" if not state.body_started else "body"
        table_type = "artifact_table"

    logical_page_label = logical_page_labels[0] if logical_page_labels else None
    return semantic_role, table_type, page_type, logical_page_label, cell_roles


def _classify_ocr_error(exc: Exception) -> str:
    if isinstance(exc, OcrConfigurationError):
        return "provider_configuration_error"
    if isinstance(exc, OcrRequestError):
        return "provider_request_failed"
    if isinstance(exc, RuntimeError):
        return "provider_unavailable"
    return "provider_execution_failed"


class DocxParser(ParserAdapter):
    name = "docx-native"

    def __init__(self, *, media_types: Sequence[str], extensions: Sequence[str]) -> None:
        self._media_types = {item.lower() for item in media_types}
        self._extensions = {item.lower() for item in extensions}

    def supports(self, *, media_type: str | None, suffix: str) -> bool:
        normalized_type = (media_type or "").lower()
        normalized_suffix = suffix.lower()
        return normalized_type in self._media_types or normalized_suffix in self._extensions

    def parse(self, request: ParseRequest) -> Sequence[Block]:
        document_path = Path(request.file_path)
        with zipfile.ZipFile(document_path) as archive:
            document_xml = archive.read("word/document.xml")

        root = ET.fromstring(document_xml)
        namespaces = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        body = root.find("./w:body", namespaces)
        blocks: list[Block] = [
            Block(
                block_id=f"{request.doc_id}-title",
                doc_id=request.doc_id,
                type=BlockType.TITLE,
                content=document_path.stem,
                metadata={
                    "page": 1,
                    "page_type": "body",
                    "parser": self.name,
                    "kind": BlockType.TITLE.value,
                    "semantic_role": SemanticRole.TITLE.value,
                },
            )
        ]
        if body is None:
            return tuple(blocks)

        state = _DocxParseState()
        position = 1
        logical_page_index = 1
        for child in list(body):
            local_name = _xml_local_name(child.tag)
            if local_name == "p":
                info = _docx_paragraph_info(child, namespaces=namespaces)
                # Advance logical page on explicit page/column breaks or
                # top-level headings (outline level 1 or 2).
                if info.has_page_break:
                    logical_page_index += 1
                elif (
                    info.outline_level is not None
                    and info.outline_level <= 2
                    and logical_page_index > 1
                ):
                    logical_page_index += 1
                if not info.text:
                    continue
                classified = _classify_docx_paragraph(info, state=state)
                if not classified.content:
                    continue
                metadata = {
                    "page": 1,
                    "logical_page": logical_page_index,
                    "page_type": classified.page_type,
                    "parser": self.name,
                    "position": position,
                    "semantic_role": classified.semantic_role,
                    "kind": classified.kind or classified.block_type.value,
                }
                if classified.normalized_title:
                    metadata["normalized_title"] = classified.normalized_title
                if classified.logical_page_label:
                    metadata["logical_page_label"] = classified.logical_page_label
                if classified.heading_level is not None:
                    metadata["heading_level"] = int(classified.heading_level)
                if classified.block_type == BlockType.TITLE and classified.semantic_role in {
                    SemanticRole.FRONT_MATTER.value,
                    SemanticRole.BODY_SECTION.value,
                    SemanticRole.APPENDIX.value,
                }:
                    metadata["is_section_heading"] = True
                    state.last_heading_text = classified.content
                    state.last_heading_role = classified.semantic_role
                    state.last_heading_page_type = classified.page_type
                elif classified.semantic_role not in _DOCX_NOISE_ROLES:
                    state.heading_context_open = False

                blocks.append(
                    Block(
                        block_id=f"{request.doc_id}-p-{position}",
                        doc_id=request.doc_id,
                        type=classified.block_type,
                        content=classified.content,
                        metadata=metadata,
                    )
                )
                position += 1
                continue

            if local_name != "tbl":
                continue

            cells = _docx_table_cells(child, namespaces=namespaces)
            if not cells:
                continue
            semantic_role, table_type, page_type, logical_page_label, cell_roles = _classify_docx_table(cells, state=state)
            content = _render_docx_table_markdown(cells)
            if not content:
                continue
            metadata = {
                "page": 1,
                "logical_page": logical_page_index,
                "page_type": page_type,
                "parser": self.name,
                "position": position,
                "semantic_role": semantic_role,
                "kind": BlockType.TABLE.value,
                "table_type": table_type,
                "cells": [list(row) for row in cells],
                "cell_semantic_roles": cell_roles,
                "rows": len(cells),
                "cols": max((len(row) for row in cells), default=0),
                "header_rows": 1,
            }
            if state.heading_context_open and state.last_heading_text:
                metadata["table_title"] = state.last_heading_text
            if logical_page_label:
                metadata["logical_page_label"] = logical_page_label
            blocks.append(
                Block(
                    block_id=f"{request.doc_id}-p-{position}",
                    doc_id=request.doc_id,
                    type=BlockType.TABLE,
                    content=content,
                    metadata=metadata,
                )
            )
            position += 1
        return tuple(blocks)


class ExcelParser(ParserAdapter):
    name = "excel-native"

    def __init__(
        self,
        *,
        media_types: Sequence[str],
        extensions: Sequence[str],
        options: Mapping[str, Any] | None = None,
    ) -> None:
        self._media_types = {item.lower() for item in media_types}
        self._extensions = {item.lower() for item in extensions}
        self._options = dict(options or {})
        self._data_only = bool(self._options.get("data_only", False))
        self._include_hidden_sheets = bool(self._options.get("include_hidden_sheets", True))
        self._max_rows_per_sheet = max(1, int(self._options.get("max_rows_per_sheet", 5000)))
        self._max_cols_per_sheet = max(1, int(self._options.get("max_cols_per_sheet", 100)))
        self._max_metadata_cells = max(1, int(self._options.get("max_metadata_cells", 1000)))

    def supports(self, *, media_type: str | None, suffix: str) -> bool:
        normalized_type = (media_type or "").lower()
        normalized_suffix = suffix.lower()
        return normalized_type in self._media_types or normalized_suffix in self._extensions

    def parse(self, request: ParseRequest) -> Sequence[Block]:
        document_path = Path(request.file_path)
        blocks: list[Block] = [self._title_block(request=request, document_path=document_path)]
        suffix = document_path.suffix.lower()
        if suffix == ".xls" or (
            suffix not in {".xlsx", ".xlsm"}
            and (request.media_type or "").lower() == "application/vnd.ms-excel"
        ):
            blocks.extend(self._parse_xls(request=request, document_path=document_path))
            return tuple(blocks)
        blocks.extend(self._parse_xlsx(request=request, document_path=document_path))
        return tuple(blocks)

    def _title_block(self, *, request: ParseRequest, document_path: Path) -> Block:
        return Block(
            block_id=f"{request.doc_id}-title",
            doc_id=request.doc_id,
            type=BlockType.TITLE,
            content=document_path.stem,
            metadata={
                "page": 1,
                "page_type": "body",
                "parser": self.name,
                "kind": BlockType.TITLE.value,
                "semantic_role": SemanticRole.TITLE.value,
            },
        )

    def _parse_xlsx(self, *, request: ParseRequest, document_path: Path) -> list[Block]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "excel-native parser requires openpyxl; install parsecore-starter[parsers]"
            ) from exc

        merged_by_sheet: dict[str, list[tuple[str, tuple[int, int, int, int]]]] = {}
        metadata_workbook = None
        try:
            metadata_workbook = load_workbook(
                filename=document_path,
                read_only=False,
                data_only=self._data_only,
            )
            for worksheet in metadata_workbook.worksheets:
                merged_by_sheet[str(worksheet.title)] = [
                    (
                        str(merged_range),
                        (
                            int(merged_range.min_row),
                            int(merged_range.min_col),
                            int(merged_range.max_row),
                            int(merged_range.max_col),
                        ),
                    )
                    for merged_range in worksheet.merged_cells.ranges
                ]
        finally:
            if metadata_workbook is not None:
                close_metadata = getattr(metadata_workbook, "close", None)
                if callable(close_metadata):
                    close_metadata()

        workbook = load_workbook(
            filename=document_path,
            read_only=True,
            data_only=self._data_only,
        )
        blocks: list[Block] = []
        try:
            position = 1
            for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
                hidden_sheet = str(getattr(worksheet, "sheet_state", "visible")) != "visible"
                if hidden_sheet and not self._include_hidden_sheets:
                    continue

                sheet_max_row = min(int(worksheet.max_row or 1), self._max_rows_per_sheet)
                sheet_max_col = min(int(worksheet.max_column or 1), self._max_cols_per_sheet)
                raw_rows: list[list[str]] = []
                formula_cells: set[tuple[int, int]] = set()
                for row in worksheet.iter_rows(
                    min_row=1,
                    max_row=sheet_max_row,
                    min_col=1,
                    max_col=sheet_max_col,
                ):
                    values: list[str] = []
                    for cell in row:
                        value = getattr(cell, "value", None)
                        is_formula = getattr(cell, "data_type", None) == "f" or (
                            isinstance(value, str) and value.startswith("=")
                        )
                        if is_formula:
                            formula_cells.add((int(getattr(cell, "row", 0)), int(getattr(cell, "column", 0))))
                        values.append(_normalize_excel_cell_value(value))
                    raw_rows.append(values)
                appended = self._append_sheet_blocks(
                    blocks=blocks,
                    request=request,
                    sheet_name=str(worksheet.title),
                    sheet_index=sheet_index,
                    hidden_sheet=hidden_sheet,
                    raw_rows=raw_rows,
                    formula_cells=formula_cells,
                    merged_ranges=merged_by_sheet.get(str(worksheet.title), []),
                    original_max_row=int(worksheet.max_row or 0),
                    original_max_col=int(worksheet.max_column or 0),
                    position=position,
                )
                position += appended
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()
        return blocks

    def _parse_xls(self, *, request: ParseRequest, document_path: Path) -> list[Block]:
        try:
            import xlrd
        except ImportError as exc:
            raise RuntimeError(
                "excel-native parser requires xlrd for .xls files; install parsecore-starter[parsers]"
            ) from exc

        try:
            workbook = xlrd.open_workbook(document_path, formatting_info=True)
        except NotImplementedError:
            workbook = xlrd.open_workbook(document_path, formatting_info=False)
        blocks: list[Block] = []
        position = 1
        for sheet_index, worksheet in enumerate(workbook.sheets(), start=1):
            hidden_sheet = int(getattr(worksheet, "visibility", 0)) != 0
            if hidden_sheet and not self._include_hidden_sheets:
                continue
            sheet_max_row = min(int(worksheet.nrows or 1), self._max_rows_per_sheet)
            sheet_max_col = min(int(worksheet.ncols or 1), self._max_cols_per_sheet)
            raw_rows: list[list[str]] = []
            for row_index in range(sheet_max_row):
                values: list[str] = []
                for col_index in range(sheet_max_col):
                    cell = worksheet.cell(row_index, col_index)
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate_as_datetime(cell.value, workbook.datemode)
                        except (TypeError, ValueError, OverflowError):
                            value = cell.value
                    else:
                        value = cell.value
                    values.append(_normalize_excel_cell_value(value))
                raw_rows.append(values)
            merged_ranges = [
                (
                    _excel_range_ref(row_low + 1, col_low + 1, row_high, col_high),
                    (row_low + 1, col_low + 1, row_high, col_high),
                )
                for row_low, row_high, col_low, col_high in getattr(worksheet, "merged_cells", [])
            ]
            appended = self._append_sheet_blocks(
                blocks=blocks,
                request=request,
                sheet_name=str(worksheet.name),
                sheet_index=sheet_index,
                hidden_sheet=hidden_sheet,
                raw_rows=raw_rows,
                formula_cells=set(),
                merged_ranges=merged_ranges,
                original_max_row=int(worksheet.nrows or 0),
                original_max_col=int(worksheet.ncols or 0),
                position=position,
            )
            position += appended
        return blocks

    def _append_sheet_blocks(
        self,
        *,
        blocks: list[Block],
        request: ParseRequest,
        sheet_name: str,
        sheet_index: int,
        hidden_sheet: bool,
        raw_rows: Sequence[Sequence[str]],
        formula_cells: set[tuple[int, int]],
        merged_ranges: Sequence[tuple[str, tuple[int, int, int, int]]],
        original_max_row: int,
        original_max_col: int,
        position: int,
    ) -> int:
        table_regions = _split_excel_table_regions(raw_rows)
        appended = 0
        table_count = len(table_regions)
        for table_index, (region_rows, region_start_row) in enumerate(table_regions, start=1):
            candidate_rows, candidate_start_row, table_title, title_row, header_row = _detect_excel_table_layout(
                region_rows,
                start_row=region_start_row,
            )
            rows, relative_start_row, start_col = _trim_excel_rows(candidate_rows)
            if not rows:
                continue
            start_row = candidate_start_row + relative_start_row - 1
            content = _render_excel_table_markdown(rows)
            if not content:
                continue
            end_row = start_row + len(rows) - 1
            end_col = start_col + max((len(row) for row in rows), default=1) - 1
            source_start_row = title_row or start_row
            source_range = (source_start_row, start_col, end_row, end_col)
            table_range = (start_row, start_col, end_row, end_col)
            table_formula_cells = {
                cell for cell in formula_cells if _excel_range_intersects((cell[0], cell[1], cell[0], cell[1]), table_range)
            }
            merged_cells = [
                label
                for label, merged_range in merged_ranges
                if _excel_range_intersects(merged_range, source_range)
            ]
            metadata: dict[str, Any] = {
                "page": sheet_index,
                "logical_page": sheet_index,
                "page_type": "body",
                "parser": self.name,
                "position": position + appended,
                "semantic_role": SemanticRole.TABLE.value,
                "kind": BlockType.TABLE.value,
                "table_type": "spreadsheet",
                "sheet_name": sheet_name,
                "sheet_index": sheet_index,
                "sheet_table_index": table_index,
                "sheet_table_count": table_count,
                "hidden_sheet": hidden_sheet,
                "row_range": f"{start_row}:{end_row}",
                "column_range": f"{_excel_column_letter(start_col)}:{_excel_column_letter(end_col)}",
                "cell_range": _excel_range_ref(start_row, start_col, end_row, end_col),
                "source_cell_range": _excel_range_ref(source_start_row, start_col, end_row, end_col),
                "rows": len(rows),
                "cols": max((len(row) for row in rows), default=0),
                "header_rows": 1 if rows else 0,
                "header_row": header_row,
                "header_values": list(rows[0]) if rows else [],
                "has_formula": bool(table_formula_cells),
                "formula_count": len(table_formula_cells),
                "truncated": (
                    original_max_row > self._max_rows_per_sheet
                    or original_max_col > self._max_cols_per_sheet
                ),
            }
            if table_title:
                metadata["table_title"] = table_title
                metadata["title_row"] = title_row
            if merged_cells:
                metadata["merged_cells"] = merged_cells
            metadata.update(_excel_cells_metadata(rows, max_cells=self._max_metadata_cells))
            blocks.append(
                Block(
                    block_id=f"{request.doc_id}-sheet-{sheet_index}-table-{table_index}",
                    doc_id=request.doc_id,
                    type=BlockType.TABLE,
                    content=content,
                    metadata=metadata,
                )
            )
            appended += 1
        return appended


class TextParser(ParserAdapter):
    name = "text-native"

    def __init__(self, *, media_types: Sequence[str], extensions: Sequence[str]) -> None:
        self._media_types = {item.lower() for item in media_types}
        self._extensions = {item.lower() for item in extensions}

    def supports(self, *, media_type: str | None, suffix: str) -> bool:
        normalized_type = (media_type or "").lower()
        normalized_suffix = suffix.lower()
        return normalized_type in self._media_types or normalized_suffix in self._extensions

    def parse(self, request: ParseRequest) -> Sequence[Block]:
        text = Path(request.file_path).read_text(encoding="utf-8")
        paragraphs = [item.strip() for item in text.split("\n\n") if item.strip()]
        blocks: list[Block] = [
            Block(
                block_id=f"{request.doc_id}-title",
                doc_id=request.doc_id,
                type=BlockType.TITLE,
                content=Path(request.file_path).stem,
                metadata={
                    "page": 1,
                    "page_type": "body",
                    "parser": self.name,
                    "semantic_role": SemanticRole.TITLE.value,
                },
            )
        ]
        for position, paragraph in enumerate(paragraphs, start=1):
            blocks.append(
                Block(
                    block_id=f"{request.doc_id}-p-{position}",
                    doc_id=request.doc_id,
                    type=BlockType.PARAGRAPH,
                    content=paragraph,
                    metadata={
                        "page": 1,
                        "page_type": "body",
                        "parser": self.name,
                        "position": position,
                        "semantic_role": SemanticRole.PARAGRAPH.value,
                    },
                )
            )
        return tuple(blocks)


class PdfTextParser(ParserAdapter):
    name = "pdf-text"

    def __init__(
        self,
        *,
        media_types: Sequence[str],
        extensions: Sequence[str],
        options: Mapping[str, Any] | None = None,
        ocr_provider_settings: OcrProviderSettings | None = None,
        boundary_refiner: Any = None,
        semantic_refiner: Any = None,
    ) -> None:
        self._media_types = {item.lower() for item in media_types}
        self._extensions = {item.lower() for item in extensions}
        post_process = {}
        if options:
            raw = options.get("post_process")
            if isinstance(raw, Mapping):
                post_process = dict(raw)

        # B3: fidelity_profile coarse knob.  Applies before individual flags.
        # "full_fidelity" – preserve everything, disable noise filters.
        # "balanced"      – default config-driven behaviour (no override).
        # "rag_clean"     – aggressively strip headers/footers and dedup lines.
        _fidelity_profile = str(options.get("fidelity_profile", "balanced") if options else "balanced").lower()
        self._fidelity_profile = _fidelity_profile
        # A3+.1 (opt-in, default False): strip repeated header/footer lines
        # detected via cross-page text similarity. Off by default so the full
        # rendered page content remains visible to downstream consumers; flip
        # ``strip_headers_footers = true`` in parsecore.toml to enable.
        self._strip_hf_enabled = bool(post_process.get("strip_headers_footers", False))
        self._merge_short_enabled = bool(post_process.get("merge_short_blocks", True))
        self._short_block_min_length = int(post_process.get("short_block_min_length", 10))
        self._hf_threshold = float(post_process.get("hf_threshold", 0.5))
        self._hf_head_n = int(post_process.get("hf_head_n", 3))
        self._hf_tail_n = int(post_process.get("hf_tail_n", 3))
        self._hf_min_line_len = int(post_process.get("hf_min_line_len", 4))
        self._split_structural_enabled = bool(
            post_process.get("split_structural_items", True)
        )
        self._structural_min_lines = int(
            post_process.get("structural_min_lines_trigger", 10)
        )
        self._split_inline_structural_enabled = bool(
            post_process.get("split_inline_structural_items", True)
        )
        self._inline_structural_min_length = int(
            post_process.get("inline_structural_min_length_trigger", 120)
        )
        self._split_toc_enabled = bool(
            post_process.get("split_toc_entries", True)
        )
        self._toc_min_entries = int(
            post_process.get("toc_min_entries_trigger", 3)
        )
        self._merge_table_enabled = bool(
            post_process.get("merge_table_continuations", True)
        )
        self._merge_figure_caption_enabled = bool(
            post_process.get("merge_figure_captions", True)
        )
        self._merge_highlights_enabled = bool(
            post_process.get("merge_highlights_entries", True)
        )
        # A3 dual-channel: use pdfplumber for layout-aware tables + text.
        # Default off so existing pypdf-only behaviour is preserved.
        self._dual_channel_enabled = bool(
            post_process.get("dual_channel", False)
        )
        self._layout_reading_order_enabled = bool(
            post_process.get("layout_reading_order", self._dual_channel_enabled)
        )
        self._dual_table_min_rows = int(
            post_process.get("dual_table_min_rows", 2)
        )
        self._dual_table_min_cols = int(
            post_process.get("dual_table_min_cols", 2)
        )
        # A6: bad-page OCR fallback. Off by default at the library level;
        # enabled in repo config so only obviously broken PDF pages fall back
        # to OCR while healthy digital text pages stay on the native path.
        self._ocr_bad_pages_enabled = bool(post_process.get("ocr_bad_pages", False))
        self._ocr_bad_page_min_cid_tokens = int(
            post_process.get("ocr_bad_page_min_cid_tokens", 5)
        )
        self._ocr_bad_page_min_cid_char_ratio = float(
            post_process.get("ocr_bad_page_min_cid_char_ratio", 0.12)
        )
        self._ocr_render_resolution = int(post_process.get("ocr_render_resolution", 144))
        self._ocr_confidence_threshold = float(
            post_process.get("ocr_confidence_threshold", 0.5)
        )
        self._ocr_merge_line_gap_ratio = float(
            post_process.get("ocr_merge_line_gap_ratio", 1.6)
        )
        self._ocr_provider_settings = (
            ocr_provider_settings or _DEFAULT_OCR_PROVIDER_SETTINGS
        )
        self._ocr_parser: ImageOcrParser | None = None
        # A4 LLM hookup: optional boundary refiner invoked on low-confidence
        # paragraphs only. None == feature disabled (default).
        self._boundary_refiner = boundary_refiner
        # Vector hookup: optional semantic refiner injected by host product.
        # ParseCore never loads vector models itself; the host owns model
        # lifecycle and credentials.  When provided, the refiner exposes
        # ``similarity(left, right) -> float`` and is consulted before LLM
        # boundary repair to merge/split adjacent paragraphs cheaply.
        self._semantic_refiner = semantic_refiner
        self._semantic_merge_threshold = float(
            post_process.get("semantic_merge_threshold", 0.86)
        )
        self._semantic_split_threshold = float(
            post_process.get("semantic_split_threshold", 0.35)
        )
        self._llm_min_length = int(
            post_process.get("llm_refine_min_length", 600)
        )
        self._llm_min_markers = int(
            post_process.get("llm_refine_min_markers", 2)
        )
        # C2: page-level OCR cache.  Enabled by default when ocr_bad_pages is
        # active; can be opted out via ``ocr_cache = false`` in post_process.
        _ocr_cache_enabled = bool(post_process.get("ocr_cache", self._ocr_bad_pages_enabled))
        _ocr_cache_ttl = int(post_process.get("ocr_cache_ttl_days", 7)) * 86400
        self._ocr_cache: PageOcrCache = (
            get_default_cache()
            if _ocr_cache_enabled
            else PageOcrCache(cache_dir=None)
        )
        # B3 override: apply fidelity_profile coarse flags after all individual
        # flag reads so it acts as a final override when explicitly set.
        if _fidelity_profile == "rag_clean":
            self._strip_hf_enabled = True
            self._merge_short_enabled = True
        elif _fidelity_profile == "full_fidelity":
            self._strip_hf_enabled = False
            self._merge_short_enabled = False
            self._split_structural_enabled = False
            self._split_inline_structural_enabled = False
            self._split_toc_enabled = False
            self._merge_table_enabled = False
            self._merge_figure_caption_enabled = False
            self._merge_highlights_enabled = False

    def supports(self, *, media_type: str | None, suffix: str) -> bool:
        normalized_type = (media_type or "").lower()
        normalized_suffix = suffix.lower()
        return normalized_type in self._media_types or normalized_suffix in self._extensions

    def _ensure_pdf_ocr_engine(self) -> tuple[Any | None, str | None, float]:
        started = time.monotonic()
        if self._ocr_parser is None:
            self._ocr_parser = ImageOcrParser(
                media_types=["image/png", "image/jpeg"],
                extensions=[".png", ".jpg", ".jpeg"],
                options={"confidence_threshold": self._ocr_confidence_threshold},
                ocr_provider_settings=self._ocr_provider_settings,
            )
        try:
            return self._ocr_parser._ensure_engine(), None, round(time.monotonic() - started, 6)
        except Exception as exc:
            return None, _classify_ocr_error(exc), round(time.monotonic() - started, 6)

    def _maybe_recover_page_with_ocr(
        self,
        page: Any,
        table_bboxes: Sequence[tuple[float, float, float, float]],
        column_count_hint: int,
        extracted_text: str | None,
    ) -> tuple[str | None, str | None, str | None, _OcrStageTimings]:
        timings = _OcrStageTimings()
        reason = _ocr_fallback_reason_for_page(
            extracted_text or "",
            min_cid_tokens=self._ocr_bad_page_min_cid_tokens,
            min_cid_char_ratio=self._ocr_bad_page_min_cid_char_ratio,
        )
        if reason is None:
            return None, None, None, timings
        attempt_started = time.monotonic()
        engine, engine_error_reason, timings.engine_init_elapsed_s = self._ensure_pdf_ocr_engine()
        if engine is None:
            timings.total_elapsed_s = round(time.monotonic() - attempt_started, 6)
            return None, reason, engine_error_reason or "provider_unavailable", timings
        text, ocr_error_reason, extract_timings = _extract_ocr_text_from_page(
            page,
            engine=engine,
            table_bboxes=table_bboxes,
            resolution=self._ocr_render_resolution,
            confidence_threshold=self._ocr_confidence_threshold,
            column_count_hint=column_count_hint,
            merge_line_gap_ratio=self._ocr_merge_line_gap_ratio,
        )
        timings.render_elapsed_s = extract_timings.render_elapsed_s
        timings.input_prepare_elapsed_s = getattr(extract_timings, "input_prepare_elapsed_s", 0.0)
        timings.engine_exec_elapsed_s = getattr(extract_timings, "engine_exec_elapsed_s", 0.0)
        timings.call_elapsed_s = extract_timings.call_elapsed_s
        timings.provider_elapsed_s = extract_timings.provider_elapsed_s
        timings.provider_det_elapsed_s = getattr(extract_timings, "provider_det_elapsed_s", 0.0)
        timings.provider_cls_elapsed_s = getattr(extract_timings, "provider_cls_elapsed_s", 0.0)
        timings.provider_rec_elapsed_s = getattr(extract_timings, "provider_rec_elapsed_s", 0.0)
        timings.provider_crop_count = int(getattr(extract_timings, "provider_crop_count", 0) or 0)
        timings.provider_cls_rotate_positive_count = int(
            getattr(extract_timings, "provider_cls_rotate_positive_count", 0) or 0
        )
        timings.provider_cls_rotate_high_count = int(
            getattr(extract_timings, "provider_cls_rotate_high_count", 0) or 0
        )
        timings.postprocess_elapsed_s = extract_timings.postprocess_elapsed_s
        timings.total_elapsed_s = round(time.monotonic() - attempt_started, 6)
        if not text:
            return None, reason, ocr_error_reason or "empty_ocr_text", timings
        return text, reason, None, timings

    def parse(self, request: ParseRequest) -> Sequence[Block]:
        PdfReader = _load_pdf_reader()
        request_enable_ocr = _resolve_request_enable_ocr(request)
        request_dual_channel = _resolve_request_dual_channel(request)
        request_layout_reading_order = _resolve_request_layout_reading_order(request)
        fast_text_profile = _uses_pdf_fast_text_profile(request)
        effective_ocr_bad_pages_enabled = (
            self._ocr_bad_pages_enabled
            if request_enable_ocr is None
            else request_enable_ocr
        )
        if fast_text_profile and request_enable_ocr is None:
            effective_ocr_bad_pages_enabled = False
        # Resolve the effective OCR strategy label for observability.
        # "force"  – caller explicitly requested full-document OCR.
        # "auto"   – bad-page detector is active (either from config or caller).
        # "off"    – OCR is disabled.
        if request_enable_ocr is True:
            ocr_strategy = "force"
        elif effective_ocr_bad_pages_enabled:
            ocr_strategy = "auto"
        else:
            ocr_strategy = "off"
        effective_layout_reading_order_enabled = (
            self._layout_reading_order_enabled
            if request_layout_reading_order is None
            else request_layout_reading_order
        )
        if fast_text_profile and request_layout_reading_order is None:
            effective_layout_reading_order_enabled = False
        base_dual_channel_enabled = (
            self._dual_channel_enabled
            if request_dual_channel is None
            else request_dual_channel
        )
        if fast_text_profile and request_dual_channel is None:
            base_dual_channel_enabled = False
        effective_dual_channel_enabled = (
            base_dual_channel_enabled
            or effective_ocr_bad_pages_enabled
            or effective_layout_reading_order_enabled
        )

        if self._boundary_refiner is not None:
            reset = getattr(self._boundary_refiner, "reset", None)
            if callable(reset):
                reset()

        reader = PdfReader(request.file_path)
        page_texts = [page.extract_text() or "" for page in reader.pages]
        layout_pages: list[_PageLayout] = []
        if effective_dual_channel_enabled:
            ocr_page_text_fn = self._maybe_recover_page_with_ocr if effective_ocr_bad_pages_enabled else None
            layout_pages = _extract_pdfplumber_layout(
                request.file_path,
                min_rows=self._dual_table_min_rows,
                min_cols=self._dual_table_min_cols,
                layout_reading_order_enabled=effective_layout_reading_order_enabled,
                ocr_page_text_fn=ocr_page_text_fn,
                ocr_cache=self._ocr_cache,
            )
            # Replace per-page text with the table-stripped pdfplumber text so
            # the existing splitters do not see table contents twice.
            for index, layout in enumerate(layout_pages):
                if index < len(page_texts) and layout.text_without_tables is not None:
                    page_texts[index] = layout.text_without_tables
        if self._strip_hf_enabled:
            cleaned_page_texts = _strip_repeated_headers_footers(
                page_texts,
                threshold=self._hf_threshold,
                head_n=self._hf_head_n,
                tail_n=self._hf_tail_n,
                min_line_len=self._hf_min_line_len,
            )
            stripped_page_numbers: set[int] = set()
            for index, (original, cleaned) in enumerate(zip(page_texts, cleaned_page_texts), start=1):
                if original.strip() and original.strip() != cleaned.strip():
                    stripped_page_numbers.add(index)
        else:
            cleaned_page_texts = list(page_texts)
            stripped_page_numbers = set()

        blocks: list[Block] = [
            Block(
                block_id=f"{request.doc_id}-title",
                doc_id=request.doc_id,
                type=BlockType.TITLE,
                content=Path(request.file_path).stem,
                metadata={
                    "page": 1,
                    "page_type": "body",
                    "parser": self.name,
                    "semantic_role": SemanticRole.TITLE.value,
                    "ocr_strategy": ocr_strategy,
                },
            )
        ]
        if fast_text_profile:
            blocks[0].metadata["profile_fast_text_path"] = True
        position = 1
        for page_number, page_text in enumerate(cleaned_page_texts, start=1):
            page_layout = (
                layout_pages[page_number - 1]
                if effective_dual_channel_enabled and page_number - 1 < len(layout_pages)
                else None
            )
            paragraphs = _split_pdf_page_text(page_text)
            if self._split_structural_enabled:
                paragraphs = _split_structural_items(
                    paragraphs, min_lines_trigger=self._structural_min_lines
                )
            if self._split_inline_structural_enabled:
                paragraphs = _split_inline_structural_items(
                    paragraphs,
                    min_length_trigger=self._inline_structural_min_length,
                )
            if self._split_toc_enabled:
                paragraphs = _split_toc_entries(
                    paragraphs, min_entries_trigger=self._toc_min_entries
                )
            if self._merge_figure_caption_enabled:
                paragraphs = _merge_figure_caption_paragraphs(paragraphs)
            if self._merge_short_enabled:
                paragraphs = _merge_short_blocks(
                    paragraphs, min_length=self._short_block_min_length
                )
            if self._merge_table_enabled:
                paragraphs = _merge_table_continuations(paragraphs)
            if self._merge_highlights_enabled:
                paragraphs = _merge_highlights_entries(paragraphs)
            if self._semantic_refiner is not None:
                paragraphs = _refine_with_semantic_similarity(
                    paragraphs,
                    refiner=self._semantic_refiner,
                    merge_threshold=self._semantic_merge_threshold,
                    split_threshold=self._semantic_split_threshold,
                )
            if self._boundary_refiner is not None:
                paragraphs = _refine_low_confidence_paragraphs(
                    paragraphs,
                    refiner=self._boundary_refiner,
                    min_length=self._llm_min_length,
                    min_markers=self._llm_min_markers,
                )
            is_highlights_page = _is_highlights_page(paragraphs)
            paragraph_roles = [
                _infer_semantic_role(paragraph, is_highlights_page=is_highlights_page)
                for paragraph in paragraphs
            ]
            page_type = _infer_page_type(
                page_number=page_number,
                roles=paragraph_roles,
                full_text="\n\n".join(paragraphs),
                has_title=(page_number == 1),
            )
            sequence = _build_page_content_sequence(paragraphs, page_layout=page_layout)
            if not sequence:
                continue
            for kind, item_index in sequence:
                if kind == "table":
                    if page_layout is None:
                        continue
                    table = page_layout.tables[item_index]
                    blocks.append(
                        Block(
                            block_id=f"{request.doc_id}-t-{position}",
                            doc_id=request.doc_id,
                            type=BlockType.TABLE,
                            content=table.render_text(),
                            metadata={
                                "page": page_number,
                                "page_type": page_type,
                                "parser": self.name,
                                "position": position,
                                "kind": "table",
                                "semantic_role": SemanticRole.TABLE.value,
                                "rows": table.row_count,
                                "cols": table.col_count,
                                "bbox": table.bbox,
                                "cells": table.cells,
                                "table_index": item_index + 1,
                            },
                        )
                    )
                    _attach_page_layout_metadata(blocks[-1].metadata, page_layout)
                    position += 1
                    continue
                if kind == "image":
                    if page_layout is None:
                        continue
                    figure = page_layout.figure_regions[item_index]
                    caption_confidence = getattr(figure, "caption_confidence", None)
                    figure_kind = str(getattr(figure, "figure_kind", "") or "").strip()
                    metadata = {
                        "page": page_number,
                        "page_type": page_type,
                        "parser": self.name,
                        "position": position,
                        "kind": "image",
                        "semantic_role": SemanticRole.IMAGE.value,
                        "bbox": figure.bbox,
                        "source_kind": figure.source_kind,
                        "figure_index": item_index + 1,
                    }
                    if isinstance(caption_confidence, (int, float)):
                        metadata["caption_confidence"] = float(caption_confidence)
                    if figure_kind:
                        metadata["figure_kind"] = figure_kind
                    if figure.object_name:
                        metadata["object_name"] = figure.object_name
                    _attach_page_layout_metadata(metadata, page_layout)
                    blocks.append(
                        Block(
                            block_id=f"{request.doc_id}-i-{position}",
                            doc_id=request.doc_id,
                            type=BlockType.IMAGE,
                            content=figure.description,
                            metadata=metadata,
                        )
                    )
                    position += 1
                    continue

                paragraph = paragraphs[item_index]
                semantic_role = paragraph_roles[item_index]
                metadata: dict[str, Any] = {
                    "page": page_number,
                    "page_type": page_type,
                    "parser": self.name,
                    "position": position,
                    "page_position": item_index + 1,
                    "semantic_role": semantic_role,
                }
                if page_layout is not None:
                    _attach_page_layout_metadata(metadata, page_layout)
                if fast_text_profile:
                    metadata["profile_fast_text_path"] = True
                if page_number in stripped_page_numbers:
                    metadata["header_footer_stripped"] = True
                blocks.append(
                    Block(
                        block_id=f"{request.doc_id}-p-{position}",
                        doc_id=request.doc_id,
                        type=BlockType.PARAGRAPH,
                        content=paragraph,
                        metadata=metadata,
                    )
                )
                position += 1
        if len(blocks) == 1:
            raise RuntimeError("No extractable text found in PDF")
        return tuple(blocks)


def _split_pdf_page_text(text: str) -> list[str]:
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    paragraphs: list[str] = []
    current_lines: list[str] = []
    for raw_line in normalized.split("\n"):
        if not raw_line.strip():
            if current_lines:
                paragraphs.append("\n".join(current_lines))
                current_lines = []
            continue
        current_lines.append(raw_line)
    if current_lines:
        paragraphs.append("\n".join(current_lines))
    return paragraphs


def _build_page_content_sequence(
    paragraphs: Sequence[str],
    *,
    page_layout: _PageLayout | None,
) -> list[tuple[str, int]]:
    paragraph_count = len(paragraphs)
    tables = list(getattr(page_layout, "tables", []) or [])
    figure_regions = list(getattr(page_layout, "figure_regions", []) or [])
    if page_layout is None or (not tables and not figure_regions):
        return [("paragraph", index) for index in range(paragraph_count)]
    if paragraph_count == 0:
        items = [
            ("image", index, float(region.bbox[1]) if len(region.bbox) >= 2 else 0.0)
            for index, region in enumerate(figure_regions)
        ] + [
            ("table", index, float(table.bbox[1]) if len(table.bbox) >= 2 else 0.0)
            for index, table in enumerate(tables)
        ]
        items.sort(key=lambda item: (item[2], item[0], item[1]))
        return [(kind, index) for kind, index, _top in items]

    height = float(page_layout.height or 0.0)
    anchored_items: list[tuple[int, float, str, int]] = []
    for table_index, table in enumerate(tables):
        top = float(table.bbox[1]) if len(table.bbox) >= 2 else 0.0
        anchor = _estimate_table_anchor_index(
            top=top,
            page_height=height,
            paragraph_count=paragraph_count,
        )
        anchored_items.append((anchor, top, "table", table_index))
    for figure_index, figure in enumerate(figure_regions):
        top = float(figure.bbox[1]) if len(figure.bbox) >= 2 else 0.0
        anchor = _estimate_table_anchor_index(
            top=top,
            page_height=height,
            paragraph_count=paragraph_count,
        )
        anchored_items.append((anchor, top, "image", figure_index))
    anchored_items.sort(key=lambda item: (item[0], item[1], item[2], item[3]))

    sequence: list[tuple[str, int]] = []
    item_cursor = 0
    for paragraph_index in range(paragraph_count + 1):
        while item_cursor < len(anchored_items) and anchored_items[item_cursor][0] == paragraph_index:
            sequence.append((anchored_items[item_cursor][2], anchored_items[item_cursor][3]))
            item_cursor += 1
        if paragraph_index < paragraph_count:
            sequence.append(("paragraph", paragraph_index))
    return sequence


def _estimate_table_anchor_index(
    *,
    top: float,
    page_height: float,
    paragraph_count: int,
) -> int:
    if paragraph_count <= 0:
        return 0
    if page_height <= 0.0:
        return paragraph_count
    clamped_top = min(max(top, 0.0), page_height)
    return min(
        paragraph_count,
        max(0, int(round((clamped_top / page_height) * paragraph_count))),
    )


def _extract_pdf_figure_regions(
    page: Any,
    *,
    page_number: int,
    table_bboxes: Sequence[tuple[float, float, float, float]],
) -> list[_FigureRegion]:
    page_width = float(page.width or 0.0)
    page_height = float(page.height or 0.0)
    if page_width <= 0.0 or page_height <= 0.0:
        return []

    try:
        raw_images = list(getattr(page, "images", None) or [])
    except Exception:
        return []
    if not raw_images:
        return []

    text_lines = _extract_pdf_text_lines(page)
    figure_regions: list[_FigureRegion] = []
    for raw_image in raw_images:
        bbox = _extract_pdf_object_bbox(raw_image, page_height=page_height)
        if bbox is None:
            continue
        description = _describe_pdf_figure_region(
            bbox=bbox,
            text_lines=text_lines,
            page_number=page_number,
        )
        if not _is_meaningful_pdf_figure_bbox(
            bbox,
            page_width=page_width,
            page_height=page_height,
            table_bboxes=table_bboxes,
            caption_confidence=description.caption_confidence,
        ):
            continue
        object_name = str(raw_image.get("name") or "").strip() or None
        figure_regions.append(
            _FigureRegion(
                bbox=bbox,
                description=description.text,
                source_kind="pdf-image",
                object_name=object_name,
                caption_confidence=description.caption_confidence,
                figure_kind=description.figure_kind,
            )
        )
    return figure_regions


def _extract_pdf_object_bbox(
    obj: Mapping[str, Any],
    *,
    page_height: float,
) -> tuple[float, float, float, float] | None:
    x0 = _coerce_float(obj.get("x0"))
    x1 = _coerce_float(obj.get("x1"))
    top = _coerce_float(obj.get("top"))
    bottom = _coerce_float(obj.get("bottom"))
    if top is None or bottom is None:
        y0 = _coerce_float(obj.get("y0"))
        y1 = _coerce_float(obj.get("y1"))
        if y0 is None or y1 is None:
            return None
        top = page_height - y1
        bottom = page_height - y0
    if x0 is None or x1 is None:
        return None
    left = min(x0, x1)
    right = max(x0, x1)
    upper = min(top, bottom)
    lower = max(top, bottom)
    if right - left <= 1.0 or lower - upper <= 1.0:
        return None
    return (left, upper, right, lower)


def _coerce_float(value: Any) -> float | None:
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    if not (result == result):
        return None
    return result


def _is_meaningful_pdf_figure_bbox(
    bbox: tuple[float, float, float, float],
    *,
    page_width: float,
    page_height: float,
    table_bboxes: Sequence[tuple[float, float, float, float]],
    caption_confidence: float = 0.0,
) -> bool:
    width = bbox[2] - bbox[0]
    height = bbox[3] - bbox[1]
    if width < 18.0 or height < 18.0:
        return False
    if page_width <= 0.0 or page_height <= 0.0:
        return False
    area_ratio = _bbox_area(bbox) / max(page_width * page_height, 1.0)
    if area_ratio >= 0.8 and caption_confidence < 0.8:
        return False
    if area_ratio < 0.0015 and caption_confidence < 0.8:
        return False
    if _is_pdf_header_footer_noise_bbox(
        bbox,
        page_width=page_width,
        page_height=page_height,
        area_ratio=area_ratio,
        caption_confidence=caption_confidence,
    ):
        return False
    for table_bbox in table_bboxes:
        if _bbox_overlap_ratio(bbox, table_bbox) >= 0.75:
            return False
    return True


def _is_pdf_header_footer_noise_bbox(
    bbox: tuple[float, float, float, float],
    *,
    page_width: float,
    page_height: float,
    area_ratio: float,
    caption_confidence: float,
) -> bool:
    if page_width <= 0.0 or page_height <= 0.0:
        return False
    if caption_confidence >= 0.8:
        return False

    top_edge = bbox[1] <= page_height * 0.12 and bbox[3] <= page_height * 0.22
    bottom_edge = bbox[3] >= page_height * 0.88 and bbox[1] >= page_height * 0.76
    negative_top_margin = bbox[1] < 0.0 and bbox[3] <= page_height * 0.24
    near_horizontal_edge = bbox[0] <= page_width * 0.08 or bbox[2] >= page_width * 0.92

    if negative_top_margin:
        return True
    if area_ratio <= 0.055 and (top_edge or bottom_edge):
        return True
    if area_ratio <= 0.018 and near_horizontal_edge and (bbox[1] <= page_height * 0.2 or bbox[3] >= page_height * 0.8):
        return True
    return False


def _bbox_area(bbox: tuple[float, float, float, float]) -> float:
    return max(0.0, bbox[2] - bbox[0]) * max(0.0, bbox[3] - bbox[1])


def _bbox_intersection_area(
    left: tuple[float, float, float, float],
    right: tuple[float, float, float, float],
) -> float:
    overlap_x0 = max(left[0], right[0])
    overlap_y0 = max(left[1], right[1])
    overlap_x1 = min(left[2], right[2])
    overlap_y1 = min(left[3], right[3])
    if overlap_x1 <= overlap_x0 or overlap_y1 <= overlap_y0:
        return 0.0
    return (overlap_x1 - overlap_x0) * (overlap_y1 - overlap_y0)


def _bbox_overlap_ratio(
    bbox: tuple[float, float, float, float],
    other: tuple[float, float, float, float],
) -> float:
    bbox_area = _bbox_area(bbox)
    if bbox_area <= 0.0:
        return 0.0
    return _bbox_intersection_area(bbox, other) / bbox_area


def _extract_pdf_text_lines(page: Any) -> list[_PdfTextLine]:
    try:
        words = page.extract_words() or []
    except Exception:
        return []
    if not words:
        return []

    normalized_words: list[tuple[float, float, float, float, str]] = []
    for word in words:
        text = " ".join(str(word.get("text") or "").split())
        if not text:
            continue
        x0 = _coerce_float(word.get("x0"))
        x1 = _coerce_float(word.get("x1"))
        top = _coerce_float(word.get("top"))
        bottom = _coerce_float(word.get("bottom"))
        if x0 is None or x1 is None or top is None or bottom is None:
            continue
        normalized_words.append((min(x0, x1), min(top, bottom), max(x0, x1), max(top, bottom), text))
    if not normalized_words:
        return []

    normalized_words.sort(key=lambda item: (item[1], item[0]))
    lines: list[_PdfTextLine] = []
    current_words: list[tuple[float, float, float, float, str]] = []
    current_top = 0.0
    for word in normalized_words:
        if not current_words:
            current_words = [word]
            current_top = word[1]
            continue
        if abs(word[1] - current_top) <= 3.0:
            current_words.append(word)
            current_top = min(current_top, word[1])
            continue
        lines.append(_build_pdf_text_line(current_words))
        current_words = [word]
        current_top = word[1]
    if current_words:
        lines.append(_build_pdf_text_line(current_words))
    return lines


def _build_pdf_text_line(words: Sequence[tuple[float, float, float, float, str]]) -> _PdfTextLine:
    text = " ".join(word[4] for word in words)
    return _PdfTextLine(
        text=text,
        x0=min(word[0] for word in words),
        top=min(word[1] for word in words),
        x1=max(word[2] for word in words),
        bottom=max(word[3] for word in words),
    )


def _describe_pdf_figure_region(
    *,
    bbox: tuple[float, float, float, float],
    text_lines: Sequence[_PdfTextLine],
    page_number: int,
) -> _FigureDescription:
    nearby_lines: list[tuple[int, float, int]] = []
    center_x = (bbox[0] + bbox[2]) / 2.0
    for index, line in enumerate(text_lines):
        line_text = " ".join(line.text.split())
        if not line_text:
            continue
        overlap_ratio = _horizontal_overlap_ratio(bbox[0], bbox[2], line.x0, line.x1)
        if overlap_ratio < 0.18 and not (line.x0 <= center_x <= line.x1):
            continue
        below_distance = line.top - bbox[3]
        above_distance = bbox[1] - line.bottom
        if -6.0 <= below_distance <= 60.0:
            nearby_lines.append((0, below_distance, index))
        elif 0.0 <= above_distance <= 28.0:
            nearby_lines.append((1, above_distance, index))

    keyword_matches = [
        candidate
        for candidate in nearby_lines
        if _PDF_FIGURE_KEYWORD_PATTERN.search(text_lines[candidate[2]].text)
        or _FIGURE_CAPTION_LABEL_PATTERN.match(text_lines[candidate[2]].text.strip())
    ]
    if not keyword_matches:
        text = f"第 {page_number} 页图示区域"
        return _FigureDescription(text=text, caption_confidence=0.15, figure_kind="generic")

    keyword_matches.sort(key=lambda item: (item[0], item[1], item[2]))
    line_index = keyword_matches[0][2]
    primary = " ".join(text_lines[line_index].text.split())
    if not primary:
        text = f"第 {page_number} 页图示区域"
        return _FigureDescription(text=text, caption_confidence=0.15, figure_kind="generic")

    parts = [primary]
    primary_is_label = bool(_FIGURE_CAPTION_LABEL_PATTERN.match(primary))
    appended_following = False
    if line_index + 1 < len(text_lines):
        following = text_lines[line_index + 1]
        following_text = " ".join(following.text.split())
        if (
            following_text
            and following.top - text_lines[line_index].bottom <= 16.0
            and _horizontal_overlap_ratio(bbox[0], bbox[2], following.x0, following.x1) >= 0.12
            and not _looks_heading_like(following_text)
        ):
            if primary_is_label or len(primary) <= 28:
                parts.append(following_text)
                appended_following = True

    description = " ".join(part for part in parts if part).strip()
    if description:
        description = description[:160]
        return _FigureDescription(
            text=description,
            caption_confidence=_pdf_figure_caption_confidence(
                description,
                primary_is_label=primary_is_label,
                appended_following=appended_following,
            ),
            figure_kind=_infer_pdf_figure_kind(description),
        )
    text = f"第 {page_number} 页图示区域"
    return _FigureDescription(text=text, caption_confidence=0.15, figure_kind="generic")


def _pdf_figure_caption_confidence(
    description: str,
    *,
    primary_is_label: bool,
    appended_following: bool,
) -> float:
    if _is_generic_pdf_figure_description(description):
        return 0.15
    has_keyword = bool(_PDF_FIGURE_KEYWORD_PATTERN.search(description))
    if primary_is_label and appended_following:
        return 0.95
    if primary_is_label or has_keyword:
        return 0.9
    return 0.65


def _infer_pdf_figure_kind(description: str) -> str:
    text = str(description or "")
    if not text or _is_generic_pdf_figure_description(text):
        return "generic"
    if re.search(r"(?:流程图|flow\s*chart|flowchart|workflow)", text, re.IGNORECASE):
        return "flowchart"
    if re.search(r"(?:结构图|组织机构图|组织架构|structure|organization(?:al)?\s+chart)", text, re.IGNORECASE):
        return "structure"
    if re.search(r"(?:示意图|原理图|布置图|diagram|schematic|layout)", text, re.IGNORECASE):
        return "diagram"
    if re.search(r"(?:Chart|Graph|曲线图|柱状图|折线图|图表)", text, re.IGNORECASE):
        return "chart"
    if re.search(r"(?:Photo|Image|Illustration|图片|照片|插图)", text, re.IGNORECASE):
        return "illustration"
    if re.search(r"(?:Fig\.?|Figure|图\s*[A-Za-z0-9一二三四五六七八九十.-]+)", text, re.IGNORECASE):
        return "figure"
    return "figure"


def _horizontal_overlap_ratio(left0: float, left1: float, right0: float, right1: float) -> float:
    overlap = min(left1, right1) - max(left0, right0)
    if overlap <= 0.0:
        return 0.0
    shortest = min(max(left1 - left0, 1.0), max(right1 - right0, 1.0))
    return overlap / shortest


def _filter_repeated_pdf_figure_regions(layouts: Sequence[_PageLayout]) -> None:
    repeated_signatures: dict[tuple[int, int, int, int], int] = {}
    for layout in layouts:
        for figure in layout.figure_regions:
            signature = _pdf_figure_signature(figure.bbox, page_width=layout.width, page_height=layout.height)
            if signature is None:
                continue
            repeated_signatures[signature] = repeated_signatures.get(signature, 0) + 1

    for layout in layouts:
        if not layout.figure_regions:
            continue
        filtered: list[_FigureRegion] = []
        for figure in layout.figure_regions:
            signature = _pdf_figure_signature(figure.bbox, page_width=layout.width, page_height=layout.height)
            repeat_count = repeated_signatures.get(signature, 0) if signature is not None else 0
            if _should_drop_repeated_pdf_figure_region(
                figure,
                repeat_count=repeat_count,
                page_width=layout.width,
                page_height=layout.height,
            ):
                continue
            filtered.append(figure)
        layout.figure_regions = filtered


def _should_drop_repeated_pdf_figure_region(
    figure: _FigureRegion,
    *,
    repeat_count: int,
    page_width: float,
    page_height: float,
) -> bool:
    if repeat_count < 3 or page_width <= 0.0 or page_height <= 0.0:
        return False
    caption_confidence = float(getattr(figure, "caption_confidence", 0.0) or 0.0)
    if caption_confidence >= 0.75:
        return False

    figure_kind = str(getattr(figure, "figure_kind", "") or "").strip().lower()
    if figure_kind and figure_kind not in {"generic", "image", "illustration"}:
        return False

    area_ratio = _bbox_area(figure.bbox) / max(page_width * page_height, 1.0)
    if not _is_generic_pdf_figure_description(figure.description) and area_ratio > 0.012:
        return False
    if _is_margin_figure_region(figure.bbox, page_width=page_width, page_height=page_height):
        return True
    if area_ratio <= 0.012:
        return True
    return repeat_count >= 5 and area_ratio <= 0.03


def _pdf_figure_signature(
    bbox: tuple[float, float, float, float],
    *,
    page_width: float,
    page_height: float,
) -> tuple[int, int, int, int] | None:
    if page_width <= 0.0 or page_height <= 0.0:
        return None
    width = bbox[2] - bbox[0]
    height = bbox[3] - bbox[1]
    return (
        int(round((bbox[0] / page_width) * 100)),
        int(round((bbox[1] / page_height) * 100)),
        int(round((width / page_width) * 100)),
        int(round((height / page_height) * 100)),
    )


def _is_margin_figure_region(
    bbox: tuple[float, float, float, float],
    *,
    page_width: float,
    page_height: float,
) -> bool:
    if page_width <= 0.0 or page_height <= 0.0:
        return False
    area_ratio = _bbox_area(bbox) / max(page_width * page_height, 1.0)
    if area_ratio > 0.03:
        return False
    return bbox[1] <= page_height * 0.18 or bbox[3] >= page_height * 0.82


def _is_generic_pdf_figure_description(description: str) -> bool:
    return bool(_GENERIC_PDF_FIGURE_DESCRIPTION_PATTERN.match(description.strip()))


_LEP_ENTRY_PATTERN = re.compile(r"(?:LIST OF EFFECTIVE PAGES|\bPage\s+[A-Z0-9.\-/]+)", re.IGNORECASE)


_HEADING_LINE_PATTERN = re.compile(
    r"^(?:\d+(?:\.\d+)*[.、)]|第[\d一二三四五六七八九十百]+[章节条款])\s"
)

_STRUCTURAL_ITEM_PATTERNS: tuple[re.Pattern[str], ...] = (
    re.compile(r"^\s*\([a-zA-Z]\)\s"),
    re.compile(r"^\s*\(\d+\)\s"),
    re.compile(
        r"^\s*(?:NOTE|WARNING|CAUTION|Note|Warning|Caution|注意|警告|小心)\s*[:：]"
    ),
)

_INLINE_STRUCTURAL_MARKER_PATTERN = re.compile(
    r"(?:\(\d+\)|\([A-Za-z]\))"
)


# TOC entry terminator: dot leaders (>=2 runs of dots separated by whitespace)
# followed by a page number OR a status token like "Not applicable"/"N/A"/"TBD".
# Used to split a single paragraph/line into multiple TOC entries even when
# entries share one line separated only by spaces.
_TOC_ENTRY_TERMINATOR = re.compile(
    r"(?:\.\s*){2,}\s*(?:\d+|(?:[A-Z]-?\d+|[A-Z]?\d+)(?:[-./][A-Z0-9]+)*|[IVXLCDM]{1,7}|Not\s+applicable|N/A|TBD)\b",
    re.IGNORECASE,
)

_TABLE_COLUMN_HEADER_PATTERN = re.compile(
    r"^\s*Index\s+Name\s+P/N\s+or\s+Type\s+Manufacturer\s*$",
    re.IGNORECASE,
)

_TABLE_NOTE_PATTERN = re.compile(r"^\s*NOTE\s*[:：]", re.IGNORECASE)

_FIGURE_CAPTION_LABEL_PATTERN = re.compile(
    r"^\s*(?:(?:FIG(?:URE)?\.?|ILLUSTRATION|IMAGE|PHOTO)\s*(?:NO\.?|NUMBER|#)?\s*[A-Za-z]?\d+(?:[.-]\d+)*(?:\s*[A-Za-z])?|图\s*[A-Za-z0-9一二三四五六七八九十]+(?:[.-]\d+)*(?:\s*[A-Za-z])?)\s*[:.)\]-]?\s*$",
    re.IGNORECASE,
)

_LOWERCASE_WORD_PATTERN = re.compile(r"\b[a-z]{3,}\b")

_HIGHLIGHTS_HEADER_PATTERN = re.compile(
    r"CHAPTER/Section/Page\s+Description of Change(?:\s+Check)?",
    re.IGNORECASE,
)

_HIGHLIGHTS_CHANGE_START_PATTERN = re.compile(
    r"^\s*(?:Changed|Added|Updated|Corrected|Revised|Removed|This page is identified|The pages of this Repair are identified)\b",
    re.IGNORECASE,
)

_HIGHLIGHTS_PAGE_REF_PATTERN = re.compile(r"\bPages?\s+\d", re.IGNORECASE)


def _split_toc_entries(
    paragraphs: list[str],
    *,
    min_entries_trigger: int = 3,
) -> list[str]:
    """Split paragraphs containing multiple TOC-style entries.

    A TOC entry is ``<title> <dot leaders> <page-or-status>``. Splitting is
    terminator-driven (not line-driven) so it correctly handles both:

    * multi-line paragraphs where each line is one entry joined by ``\\n``
    * single-line paragraphs where multiple entries share one physical line
      separated only by whitespace (common when pdfplumber packs a TOC page
      into one giant string).

    Non-numeric terminators ``Not applicable`` / ``N/A`` / ``TBD`` and
    non-decimal page markers (for example ``A-1`` / ``2-3`` / ``IV``) are
    also recognised because parts inventories and maintenance tables often
    use prefixed section pages.
    """
    if not paragraphs:
        return []

    result: list[str] = []
    for paragraph in paragraphs:
        matches = list(_TOC_ENTRY_TERMINATOR.finditer(paragraph))
        if len(matches) < min_entries_trigger:
            result.append(paragraph)
            continue

        segments: list[str] = []
        last_end = 0
        for match in matches:
            end = match.end()
            segment = paragraph[last_end:end].strip()
            if segment:
                segments.append(segment)
            last_end = end
        tail = paragraph[last_end:].strip()
        if tail:
            if segments:
                segments[-1] = segments[-1] + "\n" + tail
            else:
                segments.append(tail)
        result.extend(segments if segments else [paragraph])
    return result


def _is_highlights_page(paragraphs: Sequence[str]) -> bool:
    collapsed = [" ".join((paragraph or "").split()) for paragraph in paragraphs]
    return any("HIGHLIGHTS" in item.upper() for item in collapsed) and any(
        _HIGHLIGHTS_HEADER_PATTERN.search(item) for item in collapsed
    )


def _infer_semantic_role(
    paragraph: str,
    *,
    is_highlights_page: bool = False,
) -> str:
    stripped = paragraph.strip()
    if not stripped:
        return SemanticRole.PARAGRAPH.value
    upper = stripped.upper()
    if re.match(r"^(?:NOTE|注意)\s*[:：]", stripped, re.IGNORECASE):
        return SemanticRole.NOTE.value
    if re.match(r"^(?:WARNING|警告)\s*[:：]", stripped, re.IGNORECASE):
        return SemanticRole.WARNING.value
    if re.match(r"^(?:CAUTION|小心)\s*[:：]", stripped, re.IGNORECASE):
        return SemanticRole.CAUTION.value
    if _TOC_ENTRY_TERMINATOR.search(stripped):
        return SemanticRole.TOC_ENTRY.value
    if "LIST OF EFFECTIVE PAGES" in upper or _LEP_ENTRY_PATTERN.search(stripped):
        return SemanticRole.LEP_ENTRY.value
    if is_highlights_page and not _HIGHLIGHTS_HEADER_PATTERN.search(stripped):
        if _HIGHLIGHTS_CHANGE_START_PATTERN.match(stripped) or _HIGHLIGHTS_PAGE_REF_PATTERN.search(stripped):
            return SemanticRole.HIGHLIGHTS_ENTRY.value
    return SemanticRole.PARAGRAPH.value


def _infer_page_type(
    *,
    page_number: int,
    roles: Sequence[str],
    full_text: str,
    has_title: bool,
) -> str:
    role_set = set(roles)
    normalized_text = full_text.lower()
    stripped_text = full_text.strip()
    if SemanticRole.TOC_ENTRY.value in role_set or SemanticRole.LEP_ENTRY.value in role_set:
        return "toc"
    if any(token in normalized_text for token in ("signature", "signed by", "approved by", "签字", "签名", "审批")):
        return "signature"
    if any(token in normalized_text for token in ("appendix", "annex", "附录")):
        return "appendix"
    if page_number == 1 and has_title and not stripped_text:
        return "cover"
    return "body"


def _resolve_request_enable_ocr(request: ParseRequest) -> bool | None:
    """Return the request-level OCR override.

    Three values are possible:
    - ``True``  – caller requested full-document OCR (``enable_ocr=true``).
    - ``False`` – caller explicitly disabled OCR (``enable_ocr=false``).
    - ``None``  – no override; use the parser's built-in bad-page detection
                  (equivalent to ``enable_ocr="auto"`` or omitted).
    """
    if "enable_ocr" not in request.options:
        return None
    value = request.options.get("enable_ocr")
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    # "auto" or any unrecognised string → defer to config-level bad-page detection
    if normalized == "auto":
        return None
    return bool(value)


def _resolve_request_dual_channel(request: ParseRequest) -> bool | None:
    post_process = request.options.get("post_process")
    if isinstance(post_process, Mapping) and "dual_channel" in post_process:
        return _coerce_optional_bool(post_process.get("dual_channel"))
    return None


def _resolve_request_layout_reading_order(request: ParseRequest) -> bool | None:
    post_process = request.options.get("post_process")
    if isinstance(post_process, Mapping) and "layout_reading_order" in post_process:
        return _coerce_optional_bool(post_process.get("layout_reading_order"))

    enrichment = request.options.get("enrichment")
    if isinstance(enrichment, Mapping):
        layout_reading_order = enrichment.get("layout_reading_order")
        if isinstance(layout_reading_order, Mapping) and "enabled" in layout_reading_order:
            return _coerce_optional_bool(layout_reading_order.get("enabled"))
    return None


def _uses_pdf_fast_text_profile(request: ParseRequest) -> bool:
    profile = str(request.options.get("profile") or "").strip().lower()
    return profile in {"large-pdf-catalog", "large-pdf-ledger"}


def _coerce_optional_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    normalized = str(value).strip().lower()
    if normalized in {"1", "true", "yes", "on"}:
        return True
    if normalized in {"0", "false", "no", "off"}:
        return False
    return bool(value)


def _split_structural_items(
    paragraphs: list[str],
    *,
    min_lines_trigger: int = 10,
    min_markers: int = 2,
) -> list[str]:
    """Secondary splitter for paragraphs containing numbered/labeled items.

    Some PDFs (typical maintenance manuals) emit structured lists with single
    line breaks instead of blank lines, so the blank-line splitter returns one
    giant paragraph per page. When a paragraph spans many lines and contains
    multiple structural markers (e.g. ``(a)``, ``(26)``, ``NOTE:``), split it
    at each marker boundary.
    """
    if not paragraphs:
        return []

    result: list[str] = []
    for paragraph in paragraphs:
        lines = paragraph.split("\n")
        if len(lines) < min_lines_trigger:
            result.append(paragraph)
            continue
        marker_indices = [
            idx
            for idx, line in enumerate(lines)
            if any(pattern.match(line) for pattern in _STRUCTURAL_ITEM_PATTERNS)
        ]
        if len(marker_indices) < min_markers:
            result.append(paragraph)
            continue

        segments: list[str] = []
        prev = 0
        for idx in marker_indices:
            if idx <= prev:
                continue
            segment = "\n".join(lines[prev:idx]).strip()
            if segment:
                segments.append(segment)
            prev = idx
        tail = "\n".join(lines[prev:]).strip()
        if tail:
            segments.append(tail)
        result.extend(segments if segments else [paragraph])
    return result


def _split_inline_structural_items(
    paragraphs: list[str],
    *,
    min_length_trigger: int = 120,
    min_markers: int = 2,
) -> list[str]:
    """Split long paragraphs on inline structural markers like ``(5)`` / ``(a)``.

    Some PDFs flatten nested procedures into a single paragraph even though the
    content is clearly itemized. This helper splits only long paragraphs that
    contain multiple plain structural markers, so connector references such as
    ``(2-160)`` are left untouched.
    """
    if not paragraphs:
        return []

    result: list[str] = []
    for paragraph in paragraphs:
        text = str(paragraph or "")
        if len(text) < min_length_trigger:
            result.append(paragraph)
            continue

        split_points: list[int] = []
        for match in _INLINE_STRUCTURAL_MARKER_PATTERN.finditer(text):
            start = match.start()
            if start == 0:
                continue
            if not text[start - 1].isspace():
                continue
            split_points.append(start)

        if len(split_points) < min_markers:
            result.append(paragraph)
            continue

        parts: list[str] = []
        prev = 0
        for split_index in split_points:
            part = text[prev:split_index].strip()
            if part:
                parts.append(part)
            prev = split_index
        tail = text[prev:].strip()
        if tail:
            parts.append(tail)
        result.extend(parts if len(parts) >= min_markers + 1 else [paragraph])
    return result


def _strip_repeated_headers_footers(
    pages: list[str],
    *,
    threshold: float = 0.5,
    head_n: int = 3,
    tail_n: int = 3,
    min_line_len: int = 4,
) -> list[str]:
    """Remove header/footer lines that repeat on >= threshold fraction of non-empty pages.

    Mirrors jobcard.backend.pdf_parser._strip_repeated_headers_footers to keep behaviour
    consistent across the two parsing pipelines.
    """
    if len(pages) < 3:
        return list(pages)
    non_empty_count = sum(1 for page in pages if page.strip())
    if non_empty_count < 3:
        return list(pages)
    min_count = max(2, int(non_empty_count * threshold))

    head_counter: dict[str, int] = {}
    tail_counter: dict[str, int] = {}
    for page_text in pages:
        if not page_text.strip():
            continue
        lines = [line.strip() for line in page_text.split("\n") if line.strip()]
        if not lines:
            continue
        for line in set(lines[:head_n]):
            if len(line) >= min_line_len:
                head_counter[line] = head_counter.get(line, 0) + 1
        for line in set(lines[-tail_n:]):
            if len(line) >= min_line_len:
                tail_counter[line] = tail_counter.get(line, 0) + 1

    hf_lines: set[str] = set()
    for line, count in head_counter.items():
        if count >= min_count:
            hf_lines.add(line)
    for line, count in tail_counter.items():
        if count >= min_count:
            hf_lines.add(line)
    if not hf_lines:
        return list(pages)

    cleaned: list[str] = []
    for page_text in pages:
        lines = page_text.split("\n")
        filtered = [line for line in lines if line.strip() not in hf_lines]
        cleaned.append("\n".join(filtered).strip())
    return cleaned


def _merge_short_blocks(
    paragraphs: list[str],
    *,
    min_length: int = 10,
) -> list[str]:
    """Merge blocks shorter than ``min_length`` characters into an adjacent block.

    Heading-like tokens (numbered or all-caps short lines) are preserved as-is because
    they carry semantic structure. Other very short fragments usually come from PDF
    extraction artefacts (stray numbers, isolated letters) and inflate the very-short
    block ratio without contributing information.
    """
    if len(paragraphs) <= 1:
        return list(paragraphs)

    merged: list[str] = []
    for paragraph in paragraphs:
        stripped = paragraph.strip()
        if not stripped:
            continue
        if len(stripped) >= min_length or _looks_heading_like(stripped):
            merged.append(paragraph)
            continue
        if merged:
            merged[-1] = merged[-1] + "\n" + paragraph
        else:
            merged.append(paragraph)
    return merged


def _merge_figure_caption_paragraphs(paragraphs: list[str]) -> list[str]:
    """Keep figure labels adjacent to their caption body.

    Some PDF extraction paths emit a figure label (for example ``Figure 3-1.``)
    as a standalone paragraph and move the caption text to the next paragraph.
    Merge only this narrow pattern to avoid perturbing regular narrative flow.
    """

    if len(paragraphs) <= 1:
        return list(paragraphs)

    merged: list[str] = []
    index = 0
    while index < len(paragraphs):
        current = paragraphs[index]
        current_stripped = current.strip()
        if (
            index + 1 < len(paragraphs)
            and _FIGURE_CAPTION_LABEL_PATTERN.match(current_stripped)
        ):
            following = paragraphs[index + 1]
            following_stripped = following.strip()
            following_first_line = following_stripped.splitlines()[0] if following_stripped else ""
            if (
                following_stripped
                and not _FIGURE_CAPTION_LABEL_PATTERN.match(following_stripped)
                and not _looks_heading_like(" ".join(following_stripped.split()))
                and not _TOC_ENTRY_TERMINATOR.search(following_stripped)
                and not _TABLE_COLUMN_HEADER_PATTERN.match(following_first_line)
                and not _TABLE_NOTE_PATTERN.match(following_first_line)
                and not any(pattern.match(following_first_line) for pattern in _STRUCTURAL_ITEM_PATTERNS)
            ):
                merged.append(current.rstrip() + "\n" + following.lstrip())
                index += 2
                continue
        merged.append(current)
        index += 1
    return merged


def _merge_table_continuations(paragraphs: list[str]) -> list[str]:
    """Merge continuation fragments on tabular pages back into their owning row.

    pypdf sometimes emits one logical table row as several paragraphs: a row-start
    block containing the item description, then one or more manufacturer/address
    blocks. Restrict this merge to pages that explicitly carry the canonical tools
    table column header so we do not accidentally collapse ordinary narrative text.
    """
    if len(paragraphs) <= 4:
        return list(paragraphs)

    collapsed_paragraphs = [" ".join(paragraph.split()) for paragraph in paragraphs]
    if not any(
        _TABLE_COLUMN_HEADER_PATTERN.match(collapsed)
        for collapsed in collapsed_paragraphs
    ):
        return list(paragraphs)

    merged: list[str] = []
    seen_table_header = False
    for paragraph, collapsed in zip(paragraphs, collapsed_paragraphs, strict=False):
        stripped = paragraph.strip()
        if not stripped:
            continue
        if not seen_table_header:
            merged.append(paragraph)
            if _TABLE_COLUMN_HEADER_PATTERN.match(collapsed):
                seen_table_header = True
            continue
        if _TABLE_NOTE_PATTERN.match(stripped):
            merged.append(paragraph)
            continue
        if _looks_table_row_start(collapsed):
            merged.append(paragraph)
            continue
        if merged and not _TABLE_NOTE_PATTERN.match(merged[-1].strip()):
            merged[-1] = merged[-1] + "\n" + paragraph
        else:
            merged.append(paragraph)
    return merged


def _merge_highlights_entries(paragraphs: list[str]) -> list[str]:
    """Merge over-split entries on HIGHLIGHTS / change-log pages.

    These pages typically follow a two-column pattern: a page reference anchor
    plus one or more change-description fragments. pypdf can emit them as
    several small paragraphs (for example, ``Pages 505 and 506`` followed by
    ``Changed to ...``). Restrict this merge to explicit HIGHLIGHTS pages so
    the heuristic cannot affect normal narrative content.
    """
    if len(paragraphs) <= 2:
        return list(paragraphs)

    collapsed_paragraphs = [" ".join(paragraph.split()) for paragraph in paragraphs]
    if not any("HIGHLIGHTS" in collapsed.upper() for collapsed in collapsed_paragraphs):
        return list(paragraphs)
    if not any(_HIGHLIGHTS_HEADER_PATTERN.search(collapsed) for collapsed in collapsed_paragraphs):
        return list(paragraphs)

    merged: list[str] = []
    for paragraph, collapsed in zip(paragraphs, collapsed_paragraphs, strict=False):
        if not merged:
            merged.append(paragraph)
            continue
        previous = " ".join(merged[-1].split())
        if _should_merge_highlights_fragment(previous, collapsed):
            merged[-1] = merged[-1] + "\n" + paragraph
        else:
            merged.append(paragraph)
    return merged


def _should_merge_highlights_fragment(previous: str, current: str) -> bool:
    if _HIGHLIGHTS_HEADER_PATTERN.search(previous):
        return False
    if _HIGHLIGHTS_HEADER_PATTERN.search(current):
        return False

    current_starts_change = _HIGHLIGHTS_CHANGE_START_PATTERN.match(current) is not None
    previous_starts_change = _HIGHLIGHTS_CHANGE_START_PATTERN.match(previous) is not None
    previous_page_ref_count = _count_highlights_page_refs(previous)
    current_page_ref_count = _count_highlights_page_refs(current)
    previous_lower = previous.lstrip().lower()
    current_lower = current.lstrip().lower()

    if _looks_highlights_anchor_only(previous) and current_starts_change:
        return True
    if previous_starts_change and current_starts_change:
        if (
            previous_lower.startswith("added ")
            and current_lower.startswith("added ")
            and previous_page_ref_count == 1
            and current_page_ref_count == 1
        ):
            return True
        if current_page_ref_count == 0 and previous_page_ref_count >= 1:
            return True
    return False


def _looks_highlights_anchor_only(text: str) -> bool:
    stripped = text.strip()
    if _HIGHLIGHTS_HEADER_PATTERN.search(stripped):
        return False
    if _HIGHLIGHTS_CHANGE_START_PATTERN.match(stripped):
        return False
    return _count_highlights_page_refs(stripped) >= 1 and len(stripped) <= 80


def _count_highlights_page_refs(text: str) -> int:
    return len(_HIGHLIGHTS_PAGE_REF_PATTERN.findall(text))


def _looks_table_row_start(text: str) -> bool:
    stripped = text.strip()
    if not stripped or len(stripped) < 24:
        return False
    if _TABLE_COLUMN_HEADER_PATTERN.match(stripped) or _TABLE_NOTE_PATTERN.match(stripped):
        return False
    leading_window = stripped[:20]
    if re.search(r"\b[A-Z]\b", leading_window) is None:
        return False
    return len(_LOWERCASE_WORD_PATTERN.findall(stripped)) >= 2


def _looks_heading_like(text: str) -> bool:
    stripped = text.strip()
    if not stripped or "\n" in stripped or len(stripped) > 80:
        return False
    letters = [ch for ch in stripped if ch.isalpha()]
    if letters:
        upper_ratio = sum(1 for ch in letters if ch.isupper()) / len(letters)
        if upper_ratio >= 0.8 and len(stripped) >= 3:
            return True
    if _HEADING_LINE_PATTERN.match(stripped + " "):
        return True
    return False


def _load_pdf_reader():
    try:
        from pypdf import PdfReader  # type: ignore

        return PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore

            return PdfReader
        except ImportError as exc:
            raise RuntimeError("pypdf or PyPDF2 is required for pdf-text parser") from exc


# --- A5: image OCR parser (RapidOCR + onnxruntime) ----------------------------


class ImageOcrParser(ParserAdapter):
    """Image parser backed by RapidOCR (CPU-friendly, no GPU dependency).

    Each detected text region becomes one PARAGRAPH block whose metadata
    carries the polygon ``bbox``, recogniser ``confidence`` and reading order.
    The first emitted block is a TITLE block derived from the file stem so
    downstream chunkers and indexers behave consistently with the other
    parsers in this module.
    """

    name = "image-ocr"

    def __init__(
        self,
        *,
        media_types: Sequence[str],
        extensions: Sequence[str],
        options: Mapping[str, Any] | None = None,
        ocr_provider_settings: OcrProviderSettings | None = None,
    ) -> None:
        self._media_types = {item.lower() for item in media_types}
        self._extensions = {item.lower() for item in extensions}
        opts = dict(options or {})
        self._confidence_threshold = float(opts.get("confidence_threshold", 0.5))
        self._merge_short_enabled = bool(opts.get("merge_short_blocks", False))
        self._short_block_min_length = int(opts.get("short_block_min_length", 3))
        self._ocr_provider_settings = (
            ocr_provider_settings or _DEFAULT_OCR_PROVIDER_SETTINGS
        )
        # Cached singleton; RapidOCR initialisation loads ~50MB of ONNX models.
        self._engine: Any = None

    def supports(self, *, media_type: str | None, suffix: str) -> bool:
        normalized_type = (media_type or "").lower()
        normalized_suffix = suffix.lower()
        return normalized_type in self._media_types or normalized_suffix in self._extensions

    def _ensure_engine(self) -> Any:
        if self._engine is None:
            self._engine = build_ocr_engine(self._ocr_provider_settings)
        return self._engine

    def parse(self, request: ParseRequest) -> Sequence[Block]:
        engine = self._ensure_engine()
        document_path = Path(request.file_path)
        result, _elapse = engine(str(document_path))
        regions: list[tuple[list[list[float]], str, float]] = list(result or [])

        blocks: list[Block] = [
            Block(
                block_id=f"{request.doc_id}-title",
                doc_id=request.doc_id,
                type=BlockType.TITLE,
                content=document_path.stem,
                metadata={
                    "page": 1,
                    "parser": self.name,
                    "semantic_role": SemanticRole.TITLE.value,
                },
            )
        ]

        kept: list[tuple[list[list[float]], str, float]] = []
        for entry in regions:
            try:
                box, text, confidence = entry
            except (TypeError, ValueError):
                continue
            if not isinstance(text, str) or not text.strip():
                continue
            try:
                confidence_value = float(confidence)
            except (TypeError, ValueError):
                confidence_value = 0.0
            if confidence_value < self._confidence_threshold:
                continue
            kept.append((box, text.strip(), confidence_value))

        # Reading order: top-to-bottom, then left-to-right.
        def _order_key(item: tuple[list[list[float]], str, float]) -> tuple[float, float]:
            box = item[0]
            ys = [float(point[1]) for point in box]
            xs = [float(point[0]) for point in box]
            return (min(ys), min(xs))

        kept.sort(key=_order_key)

        position = 1
        for box, text, confidence_value in kept:
            xs = [float(point[0]) for point in box]
            ys = [float(point[1]) for point in box]
            bbox = (min(xs), min(ys), max(xs), max(ys))
            blocks.append(
                Block(
                    block_id=f"{request.doc_id}-p-{position}",
                    doc_id=request.doc_id,
                    type=BlockType.PARAGRAPH,
                    content=text,
                    metadata={
                        "page": 1,
                        "parser": self.name,
                        "position": position,
                        "bbox": bbox,
                        "confidence": confidence_value,
                        "ocr_engine": "rapidocr_onnxruntime",
                        "semantic_role": SemanticRole.PARAGRAPH.value,
                    },
                )
            )
            position += 1

        if len(blocks) == 1:
            raise RuntimeError("OCR returned no text above the configured confidence threshold")
        return tuple(blocks)


# --- A3 dual-channel: pdfplumber-backed layout extraction ---------------------


from dataclasses import dataclass, field as _dc_field


@dataclass(slots=True)
class _PdfTable:
    bbox: tuple[float, float, float, float]
    cells: list[list[str]]

    @property
    def row_count(self) -> int:
        return len(self.cells)

    @property
    def col_count(self) -> int:
        return max((len(row) for row in self.cells), default=0)

    def render_text(self) -> str:
        """Render cells as a tab-separated table for human-readable storage.

        Cells are joined per row with TAB; rows joined by newline. Empty cells
        become an empty string. The structured form is preserved in metadata so
        downstream consumers can re-parse it without relying on this rendering.
        """

        lines: list[str] = []
        for row in self.cells:
            cleaned = ["" if cell is None else str(cell).strip().replace("\t", " ") for cell in row]
            lines.append("\t".join(cleaned))
        return "\n".join(lines)


@dataclass(slots=True, frozen=True)
class _FigureDescription:
    text: str
    caption_confidence: float
    figure_kind: str


@dataclass(slots=True, frozen=True)
class _FigureRegion:
    bbox: tuple[float, float, float, float]
    description: str
    source_kind: str = "pdf-image"
    object_name: str | None = None
    caption_confidence: float = 0.15
    figure_kind: str = "generic"


@dataclass(slots=True, frozen=True)
class _PdfTextLine:
    text: str
    x0: float
    top: float
    x1: float
    bottom: float


@dataclass(slots=True)
class _PageLayout:
    page_number: int
    width: float
    height: float
    text_without_tables: str | None = None
    tables: list[_PdfTable] = _dc_field(default_factory=list)
    figure_regions: list[_FigureRegion] = _dc_field(default_factory=list)
    column_count_hint: int = 1
    layout_reading_order_applied: bool = False
    layout_reading_order_strategy: str | None = None
    layout_elapsed_s: float = 0.0
    ocr_attempt_reason: str | None = None
    ocr_fallback_reason: str | None = None
    ocr_acceptance_reason: str | None = None
    ocr_rejection_reason: str | None = None
    ocr_error_reason: str | None = None
    native_text_token_count: int = 0
    final_text_token_count: int = 0
    ocr_engine_init_elapsed_s: float = 0.0
    ocr_render_elapsed_s: float = 0.0
    ocr_input_prepare_elapsed_s: float = 0.0
    ocr_engine_exec_elapsed_s: float = 0.0
    ocr_call_elapsed_s: float = 0.0
    ocr_provider_elapsed_s: float = 0.0
    ocr_provider_det_elapsed_s: float = 0.0
    ocr_provider_cls_elapsed_s: float = 0.0
    ocr_provider_rec_elapsed_s: float = 0.0
    ocr_provider_crop_count: int = 0
    ocr_provider_cls_rotate_positive_count: int = 0
    ocr_provider_cls_rotate_high_count: int = 0
    ocr_postprocess_elapsed_s: float = 0.0
    ocr_total_elapsed_s: float = 0.0


@dataclass(slots=True)
class _OcrStageTimings:
    engine_init_elapsed_s: float = 0.0
    render_elapsed_s: float = 0.0
    input_prepare_elapsed_s: float = 0.0
    engine_exec_elapsed_s: float = 0.0
    call_elapsed_s: float = 0.0
    provider_elapsed_s: float = 0.0
    provider_det_elapsed_s: float = 0.0
    provider_cls_elapsed_s: float = 0.0
    provider_rec_elapsed_s: float = 0.0
    provider_crop_count: int = 0
    provider_cls_rotate_positive_count: int = 0
    provider_cls_rotate_high_count: int = 0
    postprocess_elapsed_s: float = 0.0
    total_elapsed_s: float = 0.0


@dataclass(slots=True)
class _OcrLine:
    bbox: tuple[float, float, float, float]
    text: str
    confidence: float
    column_index: int = 0


_PDF_FIGURE_KEYWORD_PATTERN = re.compile(
    r"(?:流程图|流程|示意图|结构图|架构图|关系图|原理图|布置图|组织机构图|图示|图片|照片|Figure|Fig\.?|Illustration|Image|Photo|Chart|Diagram|Flow\s*Chart|Flowchart|Workflow|Process\s+Map|Architecture\s+Diagram)",
    re.IGNORECASE,
)

_GENERIC_PDF_FIGURE_DESCRIPTION_PATTERN = re.compile(r"^第\s*\d+\s*页图示区域$")


def _estimate_token_count(text: str | None) -> int:
    return len(_OCR_TOKEN_RE.findall(str(text or "")))


def _extract_pdfplumber_layout(
    file_path: str,
    *,
    min_rows: int,
    min_cols: int,
    layout_reading_order_enabled: bool,
    ocr_page_text_fn: Callable[
        [Any, Sequence[tuple[float, float, float, float]], int, str | None],
        tuple[str | None, str | None, str | None, _OcrStageTimings],
    ] | None = None,
    ocr_cache: "PageOcrCache | None" = None,
) -> list[_PageLayout]:
    """Extract per-page layout (tables + table-stripped text) using pdfplumber.

    Returns an empty list when pdfplumber is unavailable; callers should treat
    that as "dual channel disabled for this run" and fall through to pypdf
    behaviour without raising.
    """

    try:
        import pdfplumber  # type: ignore
    except ImportError:
        return []

    layouts: list[_PageLayout] = []
    with pdfplumber.open(file_path) as document:
        for index, page in enumerate(document.pages, start=1):
            page_started = time.monotonic()
            tables: list[_PdfTable] = []
            try:
                found = page.find_tables() or []
            except Exception:  # pragma: no cover - pdfplumber edge cases
                found = []
            table_bboxes: list[tuple[float, float, float, float]] = []
            for table in found:
                try:
                    cells = table.extract() or []
                except Exception:
                    cells = []
                cells = [
                    [(cell or "") for cell in row]
                    for row in cells
                    if any((cell or "").strip() for cell in row)
                ]
                if len(cells) < min_rows:
                    continue
                if max((len(row) for row in cells), default=0) < min_cols:
                    continue
                bbox = tuple(float(value) for value in table.bbox)
                tables.append(_PdfTable(bbox=bbox, cells=cells))  # type: ignore[arg-type]
                table_bboxes.append(bbox)  # type: ignore[arg-type]
            figure_regions = _extract_pdf_figure_regions(
                page,
                page_number=index,
                table_bboxes=table_bboxes,
            )

            column_count_hint = _estimate_column_count(page)
            text_without_tables: str | None
            layout_reading_order_applied = False
            layout_reading_order_strategy: str | None = None
            ocr_attempt_reason: str | None = None
            ocr_fallback_reason: str | None = None
            ocr_acceptance_reason: str | None = None
            ocr_rejection_reason: str | None = None
            ocr_error_reason: str | None = None
            ocr_timings = _OcrStageTimings()
            try:
                if layout_reading_order_enabled and _should_rebuild_multi_column_text(page, column_count_hint=column_count_hint):
                    text_without_tables = _extract_text_by_columns(
                        page,
                        table_bboxes,
                        column_count=column_count_hint,
                    )
                    layout_reading_order_applied = True
                    layout_reading_order_strategy = "column-reflow"
                elif table_bboxes:
                    text_without_tables = _extract_text_excluding_bboxes(page, table_bboxes)
                else:
                    text_without_tables = page.extract_text() or ""
            except Exception:
                text_without_tables = None
            layout_elapsed_s = round(time.monotonic() - page_started, 6)
            native_text_token_count = _estimate_token_count(text_without_tables)
            final_text_token_count = native_text_token_count

            if ocr_page_text_fn is not None:
                # Check page-level OCR cache before running expensive OCR.
                _cache_hit_text: str | None = None
                if ocr_cache is not None and ocr_cache.enabled:
                    _cache_hit_text = ocr_cache.get(
                        file_path=file_path,
                        page_number=index,
                        provider_tag="rapidocr",
                        options_repr="",
                    )
                if _cache_hit_text is not None:
                    text_without_tables = _cache_hit_text
                    ocr_attempt_reason = "cid_dense"
                    ocr_fallback_reason = "cid_dense"
                    ocr_acceptance_reason = "cache_hit"
                else:
                    recovered_text, ocr_attempt_reason, ocr_error_reason, ocr_timings = ocr_page_text_fn(
                        page,
                        table_bboxes,
                        column_count_hint,
                        text_without_tables,
                    )
                    if recovered_text:
                        text_without_tables = recovered_text
                        ocr_fallback_reason = ocr_attempt_reason
                        ocr_acceptance_reason = "fallback_applied"
                        # Persist to cache for future runs.
                        if ocr_cache is not None and ocr_cache.enabled:
                            ocr_cache.put(
                                file_path=file_path,
                                page_number=index,
                                provider_tag="rapidocr",
                                text=recovered_text,
                                options_repr="",
                            )
                    elif ocr_attempt_reason is not None:
                        ocr_rejection_reason = ocr_error_reason or "quality_not_improved"

            final_text_token_count = _estimate_token_count(text_without_tables)

            layouts.append(
                _PageLayout(
                    page_number=index,
                    width=float(page.width or 0.0),
                    height=float(page.height or 0.0),
                    text_without_tables=text_without_tables,
                    tables=tables,
                    figure_regions=figure_regions,
                    column_count_hint=column_count_hint,
                    layout_reading_order_applied=layout_reading_order_applied,
                    layout_reading_order_strategy=layout_reading_order_strategy,
                    layout_elapsed_s=layout_elapsed_s,
                    ocr_attempt_reason=ocr_attempt_reason,
                    ocr_fallback_reason=ocr_fallback_reason,
                    ocr_acceptance_reason=ocr_acceptance_reason,
                    ocr_rejection_reason=ocr_rejection_reason,
                    ocr_error_reason=ocr_error_reason,
                    native_text_token_count=native_text_token_count,
                    final_text_token_count=final_text_token_count,
                    ocr_engine_init_elapsed_s=ocr_timings.engine_init_elapsed_s,
                    ocr_render_elapsed_s=ocr_timings.render_elapsed_s,
                    ocr_input_prepare_elapsed_s=ocr_timings.input_prepare_elapsed_s,
                    ocr_engine_exec_elapsed_s=ocr_timings.engine_exec_elapsed_s,
                    ocr_call_elapsed_s=ocr_timings.call_elapsed_s,
                    ocr_provider_elapsed_s=ocr_timings.provider_elapsed_s,
                    ocr_provider_det_elapsed_s=ocr_timings.provider_det_elapsed_s,
                    ocr_provider_cls_elapsed_s=ocr_timings.provider_cls_elapsed_s,
                    ocr_provider_rec_elapsed_s=ocr_timings.provider_rec_elapsed_s,
                    ocr_provider_crop_count=ocr_timings.provider_crop_count,
                    ocr_provider_cls_rotate_positive_count=ocr_timings.provider_cls_rotate_positive_count,
                    ocr_provider_cls_rotate_high_count=ocr_timings.provider_cls_rotate_high_count,
                    ocr_postprocess_elapsed_s=ocr_timings.postprocess_elapsed_s,
                    ocr_total_elapsed_s=ocr_timings.total_elapsed_s,
                )
            )
    _filter_repeated_pdf_figure_regions(layouts)
    return layouts


def _attach_page_layout_metadata(metadata: dict[str, Any], page_layout: _PageLayout) -> None:
    metadata["page_width"] = page_layout.width
    metadata["page_height"] = page_layout.height
    metadata["layout_source"] = "pdfplumber"
    metadata["column_count_hint"] = page_layout.column_count_hint
    metadata["layout_reading_order_applied"] = page_layout.layout_reading_order_applied
    if page_layout.layout_reading_order_strategy is not None:
        metadata["layout_reading_order_strategy"] = page_layout.layout_reading_order_strategy
    metadata["layout_elapsed_s"] = page_layout.layout_elapsed_s
    if page_layout.ocr_attempt_reason is not None:
        metadata["ocr_attempted"] = True
        metadata["ocr_attempt_reason"] = page_layout.ocr_attempt_reason
    if page_layout.ocr_fallback_reason is not None:
        metadata["ocr_fallback_used"] = True
        metadata["ocr_fallback_reason"] = page_layout.ocr_fallback_reason
    _ocr_acceptance_reason = getattr(page_layout, "ocr_acceptance_reason", None)
    if _ocr_acceptance_reason is not None:
        metadata["ocr_acceptance_reason"] = _ocr_acceptance_reason
    _ocr_rejection_reason = getattr(page_layout, "ocr_rejection_reason", None)
    if _ocr_rejection_reason is not None:
        metadata["ocr_rejection_reason"] = _ocr_rejection_reason
    if page_layout.ocr_error_reason is not None:
        metadata["ocr_error_reason"] = page_layout.ocr_error_reason
    if page_layout.ocr_attempt_reason is not None and page_layout.ocr_fallback_reason is None:
        metadata["ocr_rejected"] = True
    metadata["native_text_token_count"] = int(getattr(page_layout, "native_text_token_count", 0))
    metadata["final_text_token_count"] = int(getattr(page_layout, "final_text_token_count", 0))
    if page_layout.ocr_engine_init_elapsed_s > 0.0:
        metadata["ocr_engine_init_elapsed_s"] = page_layout.ocr_engine_init_elapsed_s
    if page_layout.ocr_render_elapsed_s > 0.0:
        metadata["ocr_render_elapsed_s"] = page_layout.ocr_render_elapsed_s
    input_prepare_elapsed_s = getattr(page_layout, "ocr_input_prepare_elapsed_s", 0.0)
    if input_prepare_elapsed_s > 0.0:
        metadata["ocr_input_prepare_elapsed_s"] = input_prepare_elapsed_s
    engine_exec_elapsed_s = getattr(page_layout, "ocr_engine_exec_elapsed_s", 0.0)
    if engine_exec_elapsed_s > 0.0:
        metadata["ocr_engine_exec_elapsed_s"] = engine_exec_elapsed_s
    if page_layout.ocr_call_elapsed_s > 0.0:
        metadata["ocr_call_elapsed_s"] = page_layout.ocr_call_elapsed_s
    if page_layout.ocr_provider_elapsed_s > 0.0:
        metadata["ocr_provider_elapsed_s"] = page_layout.ocr_provider_elapsed_s
    provider_det_elapsed_s = getattr(page_layout, "ocr_provider_det_elapsed_s", 0.0)
    if provider_det_elapsed_s > 0.0:
        metadata["ocr_provider_det_elapsed_s"] = provider_det_elapsed_s
    provider_cls_elapsed_s = getattr(page_layout, "ocr_provider_cls_elapsed_s", 0.0)
    if provider_cls_elapsed_s > 0.0:
        metadata["ocr_provider_cls_elapsed_s"] = provider_cls_elapsed_s
    provider_rec_elapsed_s = getattr(page_layout, "ocr_provider_rec_elapsed_s", 0.0)
    if provider_rec_elapsed_s > 0.0:
        metadata["ocr_provider_rec_elapsed_s"] = provider_rec_elapsed_s
    provider_crop_count = int(getattr(page_layout, "ocr_provider_crop_count", 0) or 0)
    if provider_crop_count > 0:
        metadata["ocr_provider_crop_count"] = provider_crop_count
    provider_cls_rotate_positive_count = int(
        getattr(page_layout, "ocr_provider_cls_rotate_positive_count", 0) or 0
    )
    if provider_cls_rotate_positive_count > 0:
        metadata["ocr_provider_cls_rotate_positive_count"] = provider_cls_rotate_positive_count
    provider_cls_rotate_high_count = int(
        getattr(page_layout, "ocr_provider_cls_rotate_high_count", 0) or 0
    )
    if provider_cls_rotate_high_count > 0:
        metadata["ocr_provider_cls_rotate_high_count"] = provider_cls_rotate_high_count
    if page_layout.ocr_postprocess_elapsed_s > 0.0:
        metadata["ocr_postprocess_elapsed_s"] = page_layout.ocr_postprocess_elapsed_s
    if page_layout.ocr_total_elapsed_s > 0.0:
        metadata["ocr_total_elapsed_s"] = page_layout.ocr_total_elapsed_s


def _ocr_fallback_reason_for_page(
    text: str,
    *,
    min_cid_tokens: int,
    min_cid_char_ratio: float,
) -> str | None:
    return detect_page_garble_reason(
        text,
        min_cid_tokens=min_cid_tokens,
        min_cid_char_ratio=min_cid_char_ratio,
    )


def _normalize_ocr_provider_timings(provider_elapsed: Any) -> tuple[float, float, float, float]:
    if isinstance(provider_elapsed, Mapping):
        try:
            total_elapsed_s = float(
                provider_elapsed.get(
                    "elapsed",
                    provider_elapsed.get("elapsed_s", provider_elapsed.get("provider_elapsed_s", 0.0)),
                )
            )
        except (TypeError, ValueError):
            total_elapsed_s = 0.0
        try:
            det_elapsed_s = float(
                provider_elapsed.get("det_elapsed_s", provider_elapsed.get("provider_det_elapsed_s", 0.0))
            )
        except (TypeError, ValueError):
            det_elapsed_s = 0.0
        try:
            cls_elapsed_s = float(
                provider_elapsed.get("cls_elapsed_s", provider_elapsed.get("provider_cls_elapsed_s", 0.0))
            )
        except (TypeError, ValueError):
            cls_elapsed_s = 0.0
        try:
            rec_elapsed_s = float(
                provider_elapsed.get("rec_elapsed_s", provider_elapsed.get("provider_rec_elapsed_s", 0.0))
            )
        except (TypeError, ValueError):
            rec_elapsed_s = 0.0
        if total_elapsed_s <= 0.0:
            total_elapsed_s = round(det_elapsed_s + cls_elapsed_s + rec_elapsed_s, 6)
        return total_elapsed_s, det_elapsed_s, cls_elapsed_s, rec_elapsed_s

    if isinstance(provider_elapsed, Sequence) and not isinstance(
        provider_elapsed,
        (str, bytes, bytearray),
    ):
        values: list[float] = []
        for item in provider_elapsed:
            try:
                values.append(float(item))
            except (TypeError, ValueError):
                continue
        if values:
            det_elapsed_s = values[0] if len(values) > 0 else 0.0
            cls_elapsed_s = values[1] if len(values) > 1 else 0.0
            rec_elapsed_s = values[2] if len(values) > 2 else 0.0
            return round(sum(values), 6), det_elapsed_s, cls_elapsed_s, rec_elapsed_s

    try:
        return float(provider_elapsed), 0.0, 0.0, 0.0
    except (TypeError, ValueError):
        return 0.0, 0.0, 0.0, 0.0


def _normalize_ocr_provider_counts(provider_elapsed: Any) -> tuple[int, int, int]:
    if not isinstance(provider_elapsed, Mapping):
        return 0, 0, 0

    def _as_int(*keys: str) -> int:
        for key in keys:
            raw_value = provider_elapsed.get(key)
            try:
                return int(raw_value)
            except (TypeError, ValueError):
                continue
        return 0

    return (
        _as_int("crop_count", "provider_crop_count"),
        _as_int("cls_rotate_positive_count", "provider_cls_rotate_positive_count"),
        _as_int("cls_rotate_high_count", "provider_cls_rotate_high_count"),
    )


def _extract_ocr_text_from_page(
    page: Any,
    *,
    engine: Any,
    table_bboxes: Sequence[tuple[float, float, float, float]],
    resolution: int,
    confidence_threshold: float,
    column_count_hint: int,
    merge_line_gap_ratio: float,
) -> tuple[str | None, str | None, _OcrStageTimings]:
    timings = _OcrStageTimings()
    try:
        import numpy as np  # type: ignore
        from PIL import ImageDraw  # type: ignore
    except ImportError:
        return None, "ocr_dependencies_missing", timings

    render_started = time.monotonic()
    try:
        rendered = page.to_image(resolution=resolution).original.convert("RGB")
    except Exception:
        timings.render_elapsed_s = round(time.monotonic() - render_started, 6)
        timings.total_elapsed_s = timings.render_elapsed_s
        return None, "ocr_render_failed", timings

    if table_bboxes:
        draw = ImageDraw.Draw(rendered)
        scale_x = rendered.width / max(float(page.width or 1.0), 1.0)
        scale_y = rendered.height / max(float(page.height or 1.0), 1.0)
        for x0, top, x1, bottom in table_bboxes:
            draw.rectangle(
                (
                    x0 * scale_x,
                    top * scale_y,
                    x1 * scale_x,
                    bottom * scale_y,
                ),
                fill="white",
            )
    timings.render_elapsed_s = round(time.monotonic() - render_started, 6)

    call_started = time.monotonic()
    prepare_started = time.monotonic()
    try:
        ocr_input = _prepare_ocr_input_image(rendered, engine=engine)
        ocr_array = np.array(ocr_input)
        timings.input_prepare_elapsed_s = round(time.monotonic() - prepare_started, 6)
    except Exception:
        timings.input_prepare_elapsed_s = round(time.monotonic() - prepare_started, 6)
        timings.call_elapsed_s = round(time.monotonic() - call_started, 6)
        timings.total_elapsed_s = round(
            timings.render_elapsed_s + timings.call_elapsed_s,
            6,
        )
        return None, "ocr_input_prepare_failed", timings

    engine_started = time.monotonic()
    try:
        result, provider_elapsed = engine(ocr_array)
    except Exception as exc:
        timings.engine_exec_elapsed_s = round(time.monotonic() - engine_started, 6)
        timings.call_elapsed_s = round(time.monotonic() - call_started, 6)
        timings.total_elapsed_s = round(
            timings.render_elapsed_s + timings.call_elapsed_s,
            6,
        )
        return None, _classify_ocr_error(exc), timings
    timings.engine_exec_elapsed_s = round(time.monotonic() - engine_started, 6)
    timings.call_elapsed_s = round(time.monotonic() - call_started, 6)
    (
        timings.provider_elapsed_s,
        timings.provider_det_elapsed_s,
        timings.provider_cls_elapsed_s,
        timings.provider_rec_elapsed_s,
    ) = _normalize_ocr_provider_timings(provider_elapsed)
    (
        timings.provider_crop_count,
        timings.provider_cls_rotate_positive_count,
        timings.provider_cls_rotate_high_count,
    ) = _normalize_ocr_provider_counts(provider_elapsed)

    postprocess_started = time.monotonic()
    lines = _collect_ocr_lines(
        result,
        confidence_threshold=confidence_threshold,
        column_count_hint=column_count_hint,
        page_width=float(rendered.width),
    )
    paragraphs = _group_ocr_lines(lines, merge_line_gap_ratio=merge_line_gap_ratio)
    timings.postprocess_elapsed_s = round(time.monotonic() - postprocess_started, 6)
    timings.total_elapsed_s = round(
        timings.render_elapsed_s + timings.call_elapsed_s + timings.postprocess_elapsed_s,
        6,
    )
    if not paragraphs:
        return None, "empty_ocr_text", timings
    return "\n\n".join(paragraphs), None, timings


def _prepare_ocr_input_image(rendered: Any, *, engine: Any) -> Any:
    if getattr(engine, "_parsecore_rapidocr", False):
        return rendered.convert("L")
    module_name = getattr(type(engine), "__module__", "")
    if "rapidocr" in module_name.lower():
        return rendered.convert("L")
    return rendered


def _collect_ocr_lines(
    result: Any,
    *,
    confidence_threshold: float,
    column_count_hint: int,
    page_width: float,
) -> list[_OcrLine]:
    lines: list[_OcrLine] = []
    for entry in result or []:
        try:
            box, text, confidence = entry
        except (TypeError, ValueError):
            continue
        if not isinstance(text, str) or not text.strip():
            continue
        try:
            confidence_value = float(confidence)
        except (TypeError, ValueError):
            confidence_value = 0.0
        if confidence_value < confidence_threshold:
            continue
        try:
            xs = [float(point[0]) for point in box]
            ys = [float(point[1]) for point in box]
        except (TypeError, ValueError, IndexError):
            continue
        bbox = (min(xs), min(ys), max(xs), max(ys))
        column_index = 0
        if column_count_hint > 1 and page_width > 0.0:
            column_width = page_width / float(column_count_hint)
            center_x = (bbox[0] + bbox[2]) / 2.0
            column_index = min(
                column_count_hint - 1,
                max(0, int(center_x / max(column_width, 1.0))),
            )
        lines.append(
            _OcrLine(
                bbox=bbox,
                text=text.strip(),
                confidence=confidence_value,
                column_index=column_index,
            )
        )

    lines.sort(key=lambda item: (item.column_index, item.bbox[1], item.bbox[0]))
    return lines


def _group_ocr_lines(
    lines: Sequence[_OcrLine],
    *,
    merge_line_gap_ratio: float,
) -> list[str]:
    if not lines:
        return []

    paragraphs: list[str] = []
    current: list[str] = [lines[0].text]
    previous = lines[0]
    for line in lines[1:]:
        previous_height = max(1.0, previous.bbox[3] - previous.bbox[1])
        current_height = max(1.0, line.bbox[3] - line.bbox[1])
        vertical_gap = line.bbox[1] - previous.bbox[3]
        x_shift = abs(line.bbox[0] - previous.bbox[0])
        same_paragraph = (
            line.column_index == previous.column_index
            and 0.0 <= vertical_gap <= max(previous_height, current_height) * merge_line_gap_ratio
            and x_shift <= max(previous_height, current_height) * 6.0
        )
        if same_paragraph:
            current.append(line.text)
        else:
            paragraphs.append(" ".join(current))
            current = [line.text]
        previous = line
    paragraphs.append(" ".join(current))
    return [paragraph for paragraph in paragraphs if paragraph.strip()]


def _should_rebuild_multi_column_text(page: Any, *, column_count_hint: int) -> bool:
    if column_count_hint <= 1:
        return False
    try:
        width = float(page.width or 0.0)
        height = float(page.height or 0.0)
        return width > 0.0 and height > 0.0
    except (TypeError, ValueError):
        return False


def _extract_text_by_columns(
    page: Any,
    bboxes: Sequence[tuple[float, float, float, float]],
    *,
    column_count: int,
) -> str:
    """Extract text region-by-region to preserve multi-column reading order.

    Used only for pages already flagged as likely multi-column. Single-column
    pages continue to use the existing page-wide extraction path.
    """

    if column_count <= 1:
        return _extract_text_excluding_bboxes(page, bboxes) if bboxes else (page.extract_text() or "")

    filtered = _filter_page_excluding_bboxes(page, bboxes)
    width = float(page.width or 0.0)
    height = float(page.height or 0.0)
    if width <= 0.0 or height <= 0.0:
        return filtered.extract_text() or ""

    column_width = width / float(column_count)
    parts: list[str] = []
    for index in range(column_count):
        x0 = index * column_width
        x1 = width if index == column_count - 1 else (index + 1) * column_width
        try:
            region = filtered.crop((x0, 0.0, x1, height))
            text = (region.extract_text() or "").strip()
        except Exception:
            text = ""
        if text:
            parts.append(text)
    if parts:
        return "\n\n".join(parts)
    return filtered.extract_text() or ""


def _estimate_column_count(page: Any) -> int:
    """Return a conservative page-level column count hint from word x positions.

    This is intentionally informational only. It does not change parse order or
    splitting behaviour; it merely tags pages that look strongly two-column so
    the regression report can surface them for human inspection.
    """

    try:
        words = page.extract_words() or []
    except Exception:
        return 1

    height = float(page.height or 0.0)
    width = float(page.width or 0.0)
    if width <= 0.0 or height <= 0.0:
        return 1

    lines: list[dict[str, float | int]] = []
    for word in words:
        text = str(word.get("text") or "").strip()
        if not text:
            continue
        try:
            x0 = float(word.get("x0", 0.0))
            x1 = float(word.get("x1", x0))
            top = float(word.get("top", 0.0))
        except (TypeError, ValueError):
            continue
        if not lines or abs(top - float(lines[-1]["top"])) > 3.0:
            lines.append({"top": top, "x0": x0, "x1": x1, "words": 1})
            continue
        lines[-1]["x0"] = min(float(lines[-1]["x0"]), x0)
        lines[-1]["x1"] = max(float(lines[-1]["x1"]), x1)
        lines[-1]["words"] = int(lines[-1]["words"]) + 1

    if len(lines) < 12:
        return 1

    left_bands: set[int] = set()
    right_bands: set[int] = set()
    wide_lines = 0
    for line in lines:
        x0 = float(line["x0"])
        x1 = float(line["x1"])
        span = x1 - x0
        top = float(line["top"])
        band = min(19, max(0, int((top / height) * 20)))
        if span >= width * 0.55:
            wide_lines += 1
            continue
        if int(line["words"]) < 2 or span > width * 0.42:
            continue
        if x1 <= width * 0.48:
            left_bands.add(band)
            continue
        if x0 >= width * 0.52:
            right_bands.add(band)

    shared_bands = left_bands & right_bands
    if (
        len(shared_bands) >= 4
        and len(left_bands) >= 6
        and len(right_bands) >= 6
        and wide_lines <= max(3, len(lines) // 4)
    ):
        return 2
    return 1


def _extract_text_excluding_bboxes(
    page: Any,
    bboxes: Sequence[tuple[float, float, float, float]],
) -> str:
    """Return ``page.extract_text()`` with words inside ``bboxes`` removed.

    pdfplumber's ``filter`` API operates per-object so we drop chars whose
    centre falls inside any table bbox before re-extracting text. Works
    regardless of pdfplumber minor version because it uses the public
    ``page.filter`` callable.
    """

    filtered = _filter_page_excluding_bboxes(page, bboxes)
    return filtered.extract_text() or ""


def _filter_page_excluding_bboxes(
    page: Any,
    bboxes: Sequence[tuple[float, float, float, float]],
) -> Any:
    def _keep(obj: Mapping[str, Any]) -> bool:
        x0 = float(obj.get("x0", 0.0))
        x1 = float(obj.get("x1", 0.0))
        top = float(obj.get("top", 0.0))
        bottom = float(obj.get("bottom", 0.0))
        cx = (x0 + x1) / 2.0
        cy = (top + bottom) / 2.0
        for bx0, btop, bx1, bbottom in bboxes:
            if bx0 <= cx <= bx1 and btop <= cy <= bbottom:
                return False
        return True

    return page.filter(_keep)


# --- A4 hookup helpers --------------------------------------------------------


_LLM_REFINE_KEYWORD_PATTERN = re.compile(
    r"\b(?:NOTE|WARNING|CAUTION|SB|Service\s+Bulletin|Step|Procedure)\b",
    re.IGNORECASE,
)


def _is_low_confidence_paragraph(
    paragraph: str,
    *,
    min_length: int,
    min_markers: int,
) -> bool:
    """Conservative gate for sending a paragraph to the LLM refiner.

    Triggers only when the paragraph is long, contains multiple structural
    markers, AND mentions a procedural keyword. This intentionally biases
    towards skipping borderline cases so the per-document call cap is spent
    on the highest-value candidates.
    """

    if len(paragraph) < min_length:
        return False
    structural_markers = sum(
        1
        for pattern in _STRUCTURAL_ITEM_PATTERNS
        for _ in pattern.finditer(paragraph)
    )
    structural_markers += len(
        re.findall(r"(?:^|\s)(?:\(\d+\)|\([A-Za-z]\))", paragraph)
    )
    if structural_markers < min_markers:
        return False
    if not _LLM_REFINE_KEYWORD_PATTERN.search(paragraph):
        return False
    return True


def _refine_with_semantic_similarity(
    paragraphs: list[str],
    *,
    refiner: Any,
    merge_threshold: float,
    split_threshold: float,
) -> list[str]:
    """Apply host-injected vector similarity to merge adjacent paragraphs.

    The host-supplied refiner exposes ``similarity(left, right) -> float``.
    When two adjacent paragraphs score above ``merge_threshold`` and the join
    would not exceed a hard length cap, they are merged.  Below
    ``split_threshold`` the paragraphs are kept separate (default behaviour).

    Any exception raised by the host refiner is swallowed to preserve the
    parse pipeline; ParseCore never owns the model lifecycle.
    """

    if not paragraphs or len(paragraphs) < 2:
        return paragraphs
    similarity = getattr(refiner, "similarity", None)
    if not callable(similarity):
        return paragraphs
    merged: list[str] = [paragraphs[0]]
    for current in paragraphs[1:]:
        previous = merged[-1]
        # Skip very short fragments and tabular markers; merging those would
        # create false joins.
        if len(previous) + len(current) > 4000:
            merged.append(current)
            continue
        try:
            score = float(similarity(left=previous, right=current))
        except Exception:
            merged.append(current)
            continue
        if score >= merge_threshold:
            merged[-1] = previous.rstrip() + "\n" + current.lstrip()
        else:
            merged.append(current)
    return merged


def _refine_low_confidence_paragraphs(
    paragraphs: list[str],
    *,
    refiner: Any,
    min_length: int,
    min_markers: int,
) -> list[str]:
    if not paragraphs:
        return paragraphs
    refine = getattr(refiner, "refine_paragraph", None)
    if not callable(refine):
        return paragraphs
    refined: list[str] = []
    for paragraph in paragraphs:
        if not _is_low_confidence_paragraph(
            paragraph, min_length=min_length, min_markers=min_markers
        ):
            refined.append(paragraph)
            continue
        try:
            parts = refine(paragraph)
        except Exception:
            refined.append(paragraph)
            continue
        if not parts:
            refined.append(paragraph)
            continue
        refined.extend(parts)
    return refined


def build_parser(
    name: str,
    *,
    media_types: Sequence[str],
    extensions: Sequence[str],
    options: Mapping[str, Any] | None = None,
    ocr_provider_settings: OcrProviderSettings | None = None,
    boundary_refiner: Any = None,
    semantic_refiner: Any = None,
) -> ParserAdapter:
    normalized = name.strip().lower()
    if normalized == "docx-native":
        return DocxParser(media_types=media_types, extensions=extensions)
    if normalized == "excel-native":
        return ExcelParser(media_types=media_types, extensions=extensions, options=options)
    if normalized == "text-native":
        return TextParser(media_types=media_types, extensions=extensions)
    if normalized == "pdf-text":
        return PdfTextParser(
            media_types=media_types,
            extensions=extensions,
            options=options,
            ocr_provider_settings=ocr_provider_settings,
            boundary_refiner=boundary_refiner,
            semantic_refiner=semantic_refiner,
        )
    if normalized == "image-ocr":
        return ImageOcrParser(
            media_types=media_types,
            extensions=extensions,
            options=options,
            ocr_provider_settings=ocr_provider_settings,
        )
    return StubParser(name=name, media_types=media_types, extensions=extensions)
