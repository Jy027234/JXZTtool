from __future__ import annotations

import csv
import io
import json
import re
import sqlite3
import tempfile
from pathlib import Path
from typing import Any, Literal


ExportDataset = Literal["pages", "lines", "tables", "quality_signals", "parse_units", "records", "coverage", "reader"]
ExportFormat = Literal["jsonl", "csv", "tsv", "xlsx", "sqlite"]

EXPORT_DATASETS = {"pages", "lines", "tables", "quality_signals", "parse_units", "records", "coverage", "reader"}
EXPORT_FORMATS = {"jsonl", "csv", "tsv", "xlsx", "sqlite"}


def export_structured_projection(
    payload: dict[str, Any],
    *,
    dataset: str,
    format: str,
    as_bytes: bool = False,
    page_start: int | None = None,
    page_end: int | None = None,
    quality_signal: str | None = None,
) -> dict[str, str | bytes]:
    """Export one structured projection dataset."""
    normalized_dataset = _normalize_dataset(dataset)
    normalized_format = _normalize_format(format)
    rows = _dataset_rows(
        payload,
        normalized_dataset,
        page_start=page_start,
        page_end=page_end,
        quality_signal=quality_signal,
    )

    content = _serialize_rows(rows, dataset=normalized_dataset, format=normalized_format)
    if isinstance(content, bytes):
        exported_content: str | bytes = content
    elif as_bytes:
        exported_content = content.encode("utf-8")
    else:
        exported_content = content

    return {
        "content_type": _content_type(normalized_format),
        "filename": _filename(payload, dataset=normalized_dataset, format=normalized_format),
        "content": exported_content,
    }


def write_structured_projection(
    payload: dict[str, Any],
    *,
    dataset: str,
    format: str,
    path: str | Path,
) -> dict[str, str | int]:
    """Write one structured projection dataset directly to disk."""
    normalized_dataset = _normalize_dataset(dataset)
    normalized_format = _normalize_format(format)
    rows = _dataset_rows(payload, normalized_dataset)
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    if normalized_format == "jsonl":
        _write_jsonl(rows, target)
    elif normalized_format in {"csv", "tsv"}:
        _write_delimited(rows, target, format=normalized_format)
    elif normalized_format == "xlsx":
        _write_xlsx(rows, dataset=normalized_dataset, path=target)
    elif normalized_format == "sqlite":
        _write_sqlite(rows, dataset=normalized_dataset, path=target)
    else:
        content = _serialize_rows(rows, dataset=normalized_dataset, format=normalized_format)
        if not isinstance(content, bytes):
            content = content.encode("utf-8")
        target.write_bytes(content)

    return {
        "content_type": _content_type(normalized_format),
        "filename": _filename(payload, dataset=normalized_dataset, format=normalized_format),
        "bytes": target.stat().st_size,
    }


def _normalize_dataset(dataset: str) -> ExportDataset:
    normalized = str(dataset or "").strip().lower()
    if normalized not in EXPORT_DATASETS:
        raise ValueError("invalid_export_dataset")
    return normalized  # type: ignore[return-value]


def _normalize_format(format: str) -> ExportFormat:
    normalized = str(format or "").strip().lower()
    if normalized not in EXPORT_FORMATS:
        raise ValueError("invalid_export_format")
    return normalized  # type: ignore[return-value]


def _dataset_rows(
    payload: dict[str, Any],
    dataset: ExportDataset,
    *,
    page_start: int | None = None,
    page_end: int | None = None,
    quality_signal: str | None = None,
) -> list[dict[str, Any]]:
    rows = payload.get(dataset)
    if dataset == "coverage" and isinstance(rows, dict):
        rows = rows.get("pages")
    if dataset == "reader" and rows is None:
        rows = payload.get("blocks")
    if rows is None:
        return []
    if not isinstance(rows, list):
        raise ValueError("invalid_export_dataset_payload")

    # P5-T11: page range 筛选（适用于 pages, lines, records, parse_units）
    if page_start is not None or page_end is not None:
        rows = _filter_by_page_range(rows, page_start=page_start, page_end=page_end)

    # P5-T11: quality_signal 筛选（适用于 records, parse_units, quality_signals, reader）
    if quality_signal is not None:
        rows = _filter_by_quality_signal(rows, quality_signal=quality_signal)

    return rows


# ── P5-T11 导出筛选辅助函数 ──

def _filter_by_page_range(
    rows: list[dict[str, Any]],
    *,
    page_start: int | None = None,
    page_end: int | None = None,
) -> list[dict[str, Any]]:
    """筛选包含页码字段的行。支持 page_number / page_start+page_end 两种模式。"""
    start = page_start
    end = page_end
    if start is not None and end is not None and start > end:
        raise ValueError("invalid_page_range")
    result = []
    for row in rows:
        row_page_start = row.get("page_start") or row.get("page_number")
        row_page_end = row.get("page_end") or row_page_start
        if row_page_start is None:
            result.append(row)  # 无页码字段的行不受筛选影响
            continue
        if start is not None and row_page_end < start:
            continue
        if end is not None and row_page_start > end:
            continue
        result.append(row)
    return result


def _filter_by_quality_signal(
    rows: list[dict[str, Any]],
    *,
    quality_signal: str | None = None,
) -> list[dict[str, Any]]:
    """筛选包含 quality_signal / quality_signal_codes 字段的行。"""
    if not quality_signal:
        return rows
    result = []
    for row in rows:
        codes = row.get("quality_signal_codes") or row.get("quality_signal")
        if isinstance(codes, list) and quality_signal in codes:
            result.append(row)
        elif isinstance(codes, str) and codes == quality_signal:
            result.append(row)
        elif row.get("code") == quality_signal:
            result.append(row)
    return result

    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("invalid_export_dataset_payload")
        normalized_rows.append(row)
    return normalized_rows


def _serialize_rows(rows: list[dict[str, Any]], *, dataset: ExportDataset, format: ExportFormat) -> str | bytes:
    if format == "jsonl":
        return "\n".join(_json_dumps(row) for row in rows)
    if format == "xlsx":
        return _serialize_xlsx(rows, dataset=dataset)
    if format == "sqlite":
        return _serialize_sqlite(rows, dataset=dataset)

    delimiter = "\t" if format == "tsv" else ","
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=_fieldnames(rows), delimiter=delimiter, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({field: _cell_value(row.get(field)) for field in writer.fieldnames or []})
    return output.getvalue()


def _write_jsonl(rows: list[dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        for index, row in enumerate(rows):
            if index:
                handle.write("\n")
            handle.write(_json_dumps(row))


def _write_delimited(rows: list[dict[str, Any]], path: Path, *, format: ExportFormat) -> None:
    delimiter = "\t" if format == "tsv" else ","
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=_fieldnames(rows), delimiter=delimiter, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: _cell_value(row.get(field)) for field in writer.fieldnames or []})


def _write_xlsx(rows: list[dict[str, Any]], *, dataset: ExportDataset, path: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - optional dependency guard
        raise RuntimeError("xlsx export requires openpyxl") from exc

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=str(dataset)[:31] or "data")
    fieldnames = _fieldnames(rows)
    if fieldnames:
        worksheet.append(fieldnames)
        for row in rows:
            worksheet.append([_cell_value(row.get(field)) for field in fieldnames])
    workbook.save(str(path))


def _write_sqlite(rows: list[dict[str, Any]], *, dataset: ExportDataset, path: Path) -> None:
    fieldnames = _fieldnames(rows)
    table_name = _sqlite_identifier(str(dataset or "data"))
    path.unlink(missing_ok=True)
    conn = sqlite3.connect(path)
    try:
        conn.execute(f"CREATE TABLE {_quote_sqlite_identifier(table_name)} ({_sqlite_columns(fieldnames)})")
        if fieldnames and rows:
            placeholders = ", ".join("?" for _ in fieldnames)
            columns = ", ".join(_quote_sqlite_identifier(field) for field in fieldnames)
            sql = f"INSERT INTO {_quote_sqlite_identifier(table_name)} ({columns}) VALUES ({placeholders})"
            conn.executemany(sql, (tuple(_cell_value(row.get(field)) for field in fieldnames) for row in rows))
        conn.commit()
    finally:
        conn.close()


def _serialize_xlsx(rows: list[dict[str, Any]], *, dataset: ExportDataset) -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - optional dependency guard
        raise RuntimeError("xlsx export requires openpyxl") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = str(dataset)[:31] or "data"
    fieldnames = _fieldnames(rows)
    if fieldnames:
        worksheet.append(fieldnames)
        for row in rows:
            worksheet.append([_cell_value(row.get(field)) for field in fieldnames])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _serialize_sqlite(rows: list[dict[str, Any]], *, dataset: ExportDataset) -> bytes:
    fieldnames = _fieldnames(rows)
    table_name = _sqlite_identifier(str(dataset or "data"))
    temp_path: str | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".sqlite", delete=False) as handle:
            temp_path = handle.name
        conn = sqlite3.connect(temp_path)
        try:
            conn.execute(
                f"CREATE TABLE {_quote_sqlite_identifier(table_name)} ({_sqlite_columns(fieldnames)})"
            )
            if fieldnames and rows:
                placeholders = ", ".join("?" for _ in fieldnames)
                columns = ", ".join(_quote_sqlite_identifier(field) for field in fieldnames)
                conn.executemany(
                    f"INSERT INTO {_quote_sqlite_identifier(table_name)} ({columns}) VALUES ({placeholders})",
                    [
                        tuple(_cell_value(row.get(field)) for field in fieldnames)
                        for row in rows
                    ],
                )
            conn.commit()
        finally:
            conn.close()
        return Path(temp_path).read_bytes()
    finally:
        if temp_path:
            Path(temp_path).unlink(missing_ok=True)


def _sqlite_columns(fieldnames: list[str]) -> str:
    if not fieldnames:
        return "_empty TEXT"
    return ", ".join(f"{_quote_sqlite_identifier(field)} TEXT" for field in fieldnames)


def _sqlite_identifier(value: str) -> str:
    normalized = re.sub(r"[^A-Za-z0-9_]+", "_", str(value or "data")).strip("_")
    if not normalized:
        normalized = "data"
    if normalized[0].isdigit():
        normalized = f"_{normalized}"
    return normalized


def _quote_sqlite_identifier(value: str) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def _fieldnames(rows: list[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            field = str(key)
            if field in seen:
                continue
            seen.add(field)
            fields.append(field)
    return fields


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return _json_dumps(value)
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _content_type(format: ExportFormat) -> str:
    if format == "jsonl":
        return "application/x-ndjson; charset=utf-8"
    if format == "tsv":
        return "text/tab-separated-values; charset=utf-8"
    if format == "xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if format == "sqlite":
        return "application/vnd.sqlite3"
    return "text/csv; charset=utf-8"


def _filename(payload: dict[str, Any], *, dataset: ExportDataset, format: ExportFormat) -> str:
    doc_id = str(payload.get("doc_id") or "document")
    safe_doc_id = re.sub(r"[^A-Za-z0-9._-]+", "-", doc_id).strip("-._") or "document"
    return f"{safe_doc_id}-{dataset}.{format}"


__all__ = ["export_structured_projection", "write_structured_projection"]
